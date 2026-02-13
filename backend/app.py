import json
import logging
import os
import re
import base64
import csv
import io
import asyncio
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlparse
from uuid import uuid4
from mimetypes import guess_extension

import anyio
import requests

from dotenv import load_dotenv

from fastapi import (
    Depends,
    FastAPI,
    File,
    Form,
    Header,
    HTTPException,
    Query,
    Request,
    UploadFile,
    status,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from passlib.context import CryptContext
from pydantic import ValidationError
from sqlalchemy.exc import IntegrityError
from sqlalchemy import delete, func, or_, text
from sqlmodel import Session, select

from audio_utils import ensure_voice_note_audio
from evolution import EvolutionClient
from wppconnect import WPPConnectClient
from uazapi import UazapiClient, UazapiAdminClient
from media_storage import get_media_storage
from models import (
    AuthLogin,
    AuthRegister,
    AuthResponse,
    Conversation,
    ConversationAvatar,
    ConversationPin,
    ConversationNote,
    ConversationNoteCreate,
    ConversationNoteRead,
    ConversationCreate,
    ConversationUpdate,
    ConversationRead,
    ConversationSummary,
    DeviceSession,
    DeviceSessionStatus,
    IncomingMessagePayload,
    CallStartRequest,
    CallRejectRequest,
    CallActionResponse,
    Message,
    MessageCreate,
    MessageDirection,
    MessageType,
    MessageRead,
    MessageDeletion,
    MessageReaction,
    MessageReactRequest,
    MessageEditRequest,
    OrganizationCredentialsRead,
    OrganizationCredentialsUpdate,
    Organization,
    OrganizationRead,
    SessionRotateRequest,
    SessionToken,
    SessionRead,
    User,
    UserCreate,
    UserRead,
    UserUpdate,
    Department,
    DepartmentCreate,
    DepartmentUpdate,
    DepartmentRead,
    DepartmentUserLink,
    DepartmentUsersUpdate,
    Tag,
    TagCreate,
    TagRead,
    BulkCampaign,
    BulkCampaignRead,
    BulkCampaignCreate,
    BulkCustomFieldMapping,
    BulkContactsMapping,
    ConversationTagLink,
    Reminder,
    ReminderCreate,
    ReminderRead,
    ReminderUpdate,
    default_message_preview_length,
    CollaboratorOrganizationCreate,
    CollaboratorOrganizationRead,
    OrganizationUpdate,
)

USERNAME_PATTERN = re.compile(r"^[a-z0-9._]+$")  # validated no spaces

DOTENV_PATH = Path(__file__).resolve().parent / ".env"
if DOTENV_PATH.exists():
    load_dotenv(DOTENV_PATH)

from db import DATA_DIR, engine, get_session, init_db, IS_SQLITE

app = FastAPI(title="Chat de Valor", version="0.3.0")
evolution_client = EvolutionClient()
frontend_dir = Path(__file__).resolve().parent.parent / "frontend"
media_storage = get_media_storage(DATA_DIR)
pwd_context = CryptContext(schemes=["pbkdf2_sha256"], deprecated="auto")
DEFAULT_ORG_NAME = os.getenv("CHAT_DEFAULT_ORG_NAME", "Empresa Demo")
logger = logging.getLogger("chatdevalor")
DEFAULT_ORG_SLUG = os.getenv("CHAT_DEFAULT_ORG_SLUG", "demo")
DEFAULT_ADMIN_USERNAME = os.getenv("CHAT_ADMIN_USERNAME", "admin").lower()
DEFAULT_ADMIN_PASSWORD = os.getenv("CHAT_ADMIN_PASSWORD", "admin123")
COLLAB_USERNAME = os.getenv("CHAT_COLLAB_USERNAME", "colaborador")
COLLAB_PASSWORD = os.getenv("CHAT_COLLAB_PASSWORD", "colab1234")
COLLAB_FULL_NAME = os.getenv("CHAT_COLLAB_FULL_NAME", "Equipe Nexen")
WEBHOOK_URL = os.getenv("CHAT_WEBHOOK_URL")
WEBHOOK_EVENTS = ["messages", "connection"]
WEBHOOK_EXCLUDE_MESSAGES = ["wasSentByApi"]
UAZAPI_USE_GLOBAL_WEBHOOK = str(
    os.getenv("UAZAPI_USE_GLOBAL_WEBHOOK", "0")
).strip().lower() in {"1", "true", "yes", "on"}
logger.info("UAZAPI_USE_GLOBAL_WEBHOOK=%s", UAZAPI_USE_GLOBAL_WEBHOOK)
MEDIA_DEBUG_FILE = DATA_DIR / "media" / "download-debug.json"
WEBHOOK_LOG_DIR = DATA_DIR / "logs"
AVATAR_CACHE_TTL = timedelta(hours=12)
UAZAPI_ADMIN_BASE_URL = (os.getenv("UAZAPI_ADMIN_BASE_URL") or "").rstrip("/")
UAZAPI_ADMIN_TOKEN = os.getenv("UAZAPI_ADMIN_TOKEN", "")
UAZAPI_DEFAULT_BASE_URL = os.getenv("UAZAPI_DEFAULT_BASE_URL") or UAZAPI_ADMIN_BASE_URL
if UAZAPI_DEFAULT_BASE_URL:
    UAZAPI_DEFAULT_BASE_URL = UAZAPI_DEFAULT_BASE_URL.rstrip("/")
logger.info("UAZAPI_USE_GLOBAL_WEBHOOK=%s", UAZAPI_USE_GLOBAL_WEBHOOK)

# Realtime (SSE) broker: pub/sub em memória por organização.
# Observação: em múltiplos workers/processos, cada processo terá seu próprio broker.
SSE_QUEUE_MAXSIZE = 200
SSE_PING_INTERVAL_SECONDS = 15
_sse_lock = asyncio.Lock()
_sse_subscribers: Dict[int, List[asyncio.Queue]] = {}


async def _sse_register(org_id: int) -> asyncio.Queue:
    queue: asyncio.Queue = asyncio.Queue(maxsize=SSE_QUEUE_MAXSIZE)
    async with _sse_lock:
        _sse_subscribers.setdefault(org_id, []).append(queue)
    return queue


async def _sse_unregister(org_id: int, queue: asyncio.Queue) -> None:
    async with _sse_lock:
        items = _sse_subscribers.get(org_id) or []
        try:
            items.remove(queue)
        except ValueError:
            return
        if not items:
            _sse_subscribers.pop(org_id, None)


async def _sse_publish(org_id: int, event: str, payload: Dict[str, Any]) -> None:
    async with _sse_lock:
        queues = list(_sse_subscribers.get(org_id) or [])
    for queue in queues:
        try:
            queue.put_nowait((event, payload))
        except asyncio.QueueFull:
            continue


def sse_notify(org_id: Optional[int], event: str, payload: Dict[str, Any]) -> None:
    if not org_id:
        return
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        try:
            anyio.from_thread.run(_sse_publish, int(org_id), event, payload)
        except Exception:
            return
    else:
        loop.create_task(_sse_publish(int(org_id), event, payload))

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

if media_storage.is_local() and media_storage.local_base_path:
    app.mount(
        "/media",
        StaticFiles(directory=media_storage.local_base_path, check_dir=False),
        name="media",
    )


@app.on_event("startup")
def on_startup() -> None:
    init_db()
    ensure_database_columns()
    with Session(engine) as session:
        organization = ensure_default_organization(session)
        default_user = ensure_default_user(session, organization)
        ensure_default_collaborator(session)
        ensure_session_exists(session, organization.id, default_user.id)


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9-]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value or str(uuid4())[:8]


INSTANCE_NAME_SANITIZER = re.compile(r"[^a-z0-9_-]+")


def sanitize_instance_name(value: Optional[str]) -> str:
    if not value:
        return ""
    normalized = INSTANCE_NAME_SANITIZER.sub("", value.lower())
    return normalized.strip("-_")


def generate_uazapi_instance_name(preferred: Optional[str], fallback_seed: str) -> str:
    """
    Define o nome final da instância. Mantém o valor informado (sanitizado) ou
    gera um identificador baseado no seed + um sufixo aleatório para evitar conflitos.
    """
    candidate = sanitize_instance_name(preferred)
    if candidate:
        return candidate[:40]
    fallback = sanitize_instance_name(fallback_seed) or "instancia"
    suffix = uuid4().hex[:6]
    return f"{fallback[:20]}{suffix}"


def normalize_username(value: str) -> str:
    normalized = value.strip().lower()
    if not USERNAME_PATTERN.fullmatch(normalized):
        raise HTTPException(
            status_code=422,
            detail="Usuario deve conter apenas letras, numeros, '.' ou '_'",
        )
    return normalized


def require_admin(user: User) -> None:
    if not user.is_admin:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Requer perfil administrador")


def get_default_owner(session: Session, organization_id: int) -> User:
    owner = session.exec(
        select(User)
        .where(User.organization_id == organization_id)
        .where(User.is_admin == True)
    ).first()
    if owner:
        return owner
    owner = session.exec(
        select(User).where(User.organization_id == organization_id)
    ).first()
    if not owner:
        raise HTTPException(status_code=400, detail="Organizacao sem usuarios")
    return owner


def build_organization_read(organization: Optional[Organization]) -> Optional[OrganizationRead]:
    if not organization:
        return None
    org_read = OrganizationRead.model_validate(organization)
    org_read.uazapi_server_url = organization.uazapi_server_url
    org_read.has_uazapi_credentials = bool(
        organization.uazapi_server_url and organization.uazapi_admin_token
    )
    org_read.webhook_url = organization.webhook_url
    return org_read


def has_other_admin(session: Session, organization_id: int, exclude_user_id: Optional[int] = None) -> bool:
    query = select(User).where(User.organization_id == organization_id).where(User.is_admin == True)
    if exclude_user_id:
        query = query.where(User.id != exclude_user_id)
    return session.exec(query).first() is not None


def get_user_in_organization(
    session: Session, user_id: int, organization_id: int
) -> Optional[User]:
    user = session.get(User, user_id)
    if not user or user.organization_id != organization_id:
        return None
    return user


def ensure_conversation_access(conversation: Conversation, user: User) -> None:
    if conversation.organization_id != user.organization_id:
        raise HTTPException(status_code=404, detail="Conversa nao encontrada")
    if conversation.owner_user_id not in (user.id, None) and not user.is_admin:
        raise HTTPException(status_code=403, detail="Conversa nao pertence ao usuario")


def ensure_reply_target(
    session: Session,
    conversation_id: int,
    reply_to_message_id: Optional[int],
) -> Optional[Message]:
    if not reply_to_message_id:
        return None
    quoted_message = session.get(Message, reply_to_message_id)
    if not quoted_message or quoted_message.conversation_id != conversation_id:
        raise HTTPException(status_code=400, detail="Mensagem citada invalida")
    return quoted_message


def ensure_reminder_access(reminder: Reminder, user: User) -> None:
    if reminder.organization_id != user.organization_id:
        raise HTTPException(status_code=404, detail="Lembrete nao encontrado")
    if reminder.owner_user_id not in (user.id, None) and not user.is_admin:
        raise HTTPException(status_code=403, detail="Lembrete nao pertence ao usuario")


def ensure_default_organization(session: Session) -> Organization:
    organization = session.exec(
        select(Organization).where(Organization.slug == DEFAULT_ORG_SLUG)
    ).first()
    if organization:
        return organization
    organization = Organization(name=DEFAULT_ORG_NAME, slug=DEFAULT_ORG_SLUG)
    session.add(organization)
    session.commit()
    session.refresh(organization)
    return organization


def ensure_default_user(session: Session, organization: Organization) -> User:
    username = normalize_username(DEFAULT_ADMIN_USERNAME)
    user = session.exec(
        select(User).where(User.username == username)
    ).first()
    if user:
        if user.organization_id != organization.id:
            user.organization_id = organization.id
            user.is_admin = True
            session.add(user)
            session.commit()
            session.refresh(user)
        return user
    user = User(
        username=DEFAULT_ADMIN_USERNAME,
        full_name="Administrador",
        password_hash=get_password_hash(DEFAULT_ADMIN_PASSWORD),
        organization_id=organization.id,
        is_admin=True,
    )
    session.add(user)
    session.commit()
    session.refresh(user)
    return user


def ensure_default_collaborator(session: Session) -> User:
    username = normalize_username(COLLAB_USERNAME)
    user = session.exec(
        select(User)
        .where(User.username == username)
        .where(User.organization_id == None)  # noqa: E711
    ).first()
    if user:
        updated = False
        if not user.is_admin:
            user.is_admin = True
            updated = True
        if not verify_password(COLLAB_PASSWORD, user.password_hash):
            user.password_hash = get_password_hash(COLLAB_PASSWORD)
            updated = True
        if user.full_name != COLLAB_FULL_NAME:
            user.full_name = COLLAB_FULL_NAME
            updated = True
        if updated:
            session.add(user)
            session.commit()
            session.refresh(user)
        return user
    user = User(
        username=username,
        full_name=COLLAB_FULL_NAME,
        password_hash=get_password_hash(COLLAB_PASSWORD),
        organization_id=None,
        is_admin=True,
    )
    session.add(user)
    session.commit()
    session.refresh(user)
    return user


def ensure_database_columns() -> None:
    if not IS_SQLITE:
        return
    with engine.connect() as conn:
        org_columns = {
            row[1] for row in conn.execute(text("PRAGMA table_info('organization')")).fetchall()
        }
        if "is_active" not in org_columns:
            conn.execute(
                text("ALTER TABLE organization ADD COLUMN is_active BOOLEAN NOT NULL DEFAULT 1")
            )
        if "webhook_url" not in org_columns:
            conn.execute(
                text("ALTER TABLE organization ADD COLUMN webhook_url VARCHAR")
            )
        if "user_limit" not in org_columns:
            conn.execute(
                text("ALTER TABLE organization ADD COLUMN user_limit INTEGER NOT NULL DEFAULT 0")
            )
        session_columns = {
            row[1] for row in conn.execute(text("PRAGMA table_info('devicesession')")).fetchall()
        }
        if "webhook_id" not in session_columns:
            conn.execute(
                text("ALTER TABLE devicesession ADD COLUMN webhook_id VARCHAR")
            )


def ensure_session_exists(
    session: Session, organization_id: int, owner_user_id: Optional[int] = None
) -> DeviceSession:
    base_query = (
        select(DeviceSession)
        .where(DeviceSession.organization_id == organization_id)
        .order_by(DeviceSession.created_at.desc())
    )
    if owner_user_id is not None:
        owned_session = session.exec(
            base_query.where(DeviceSession.owner_user_id == owner_user_id)
        ).first()
        if owned_session:
            return owned_session
        shared_session = session.exec(
            base_query.where(DeviceSession.owner_user_id == None)  # noqa: E711
        ).first()
        if shared_session:
            return shared_session
    existing = session.exec(base_query).first()
    if existing:
        return existing
    new_session = DeviceSession(
        organization_id=organization_id,
        owner_user_id=owner_user_id,
        status=DeviceSessionStatus.disconnected,
        provider="uazapi",
    )
    session.add(new_session)
    session.commit()
    session.refresh(new_session)
    return new_session


def session_has_credentials(device_session: DeviceSession) -> bool:
    return all(
        [
            device_session.integration_base_url,
            device_session.integration_instance_id,
            device_session.integration_token,
        ]
    )


def ensure_voice_call_capability(device_session: DeviceSession) -> None:
    provider = ((device_session.provider or "").strip() or "uazapi").lower()
    if provider != "uazapi":
        raise HTTPException(
            status_code=400,
            detail="Chamadas de voz estao disponiveis apenas para integracoes com a Uazapi. Ajuste o provedor em Integracao > Gerar novo QR.",
        )


def ensure_admin_access(user: User) -> None:
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Acesso restrito a administradores.")


def get_organization_user_count(session: Session, organization_id: int) -> int:
    result = session.exec(
        select(func.count(User.id)).where(User.organization_id == organization_id)
    ).one()
    return int(result) if result else 0


def get_primary_admin_username(session: Session, organization_id: int) -> Optional[str]:
    admin_user = session.exec(
        select(User)
        .where(User.organization_id == organization_id)
        .where(User.is_admin == True)
        .order_by(User.created_at.asc())
    ).first()
    return admin_user.username if admin_user else None


def serialize_collaborator_organization(
    organization: Organization, user_count: int, default_admin_username: Optional[str]
) -> CollaboratorOrganizationRead:
    return CollaboratorOrganizationRead(
        id=organization.id,
        name=organization.name,
        slug=organization.slug,
        created_at=organization.created_at,
        uazapi_server_url=organization.uazapi_server_url,
        has_admin_token=bool(organization.uazapi_admin_token),
        default_admin_username=default_admin_username,
        user_count=user_count,
        is_active=organization.is_active,
        user_limit=organization.user_limit,
        webhook_url=organization.webhook_url,
    )


def normalize_whatsapp_jid(jid: Optional[str]) -> Optional[str]:
    if not jid:
        return None
    local = jid.split("@", 1)[0]
    digits = clean_phone_digits(local)
    return digits or None


def clean_phone_digits(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    digits = "".join(ch for ch in value if ch.isdigit())
    return digits or None


def phone_variants(value: Optional[str]) -> List[str]:
    digits = clean_phone_digits(value)
    if not digits:
        return []
    variants: List[str] = []
    variants.append(digits)
    if digits.startswith("55"):
        area_end = 4
        local = digits[area_end:]
        if len(local) >= 9 and local.startswith("9"):
            variants.append(digits[:area_end] + local[1:])
        if len(local) == 8:
            variants.append(digits[:area_end] + "9" + local)
    seen: Dict[str, bool] = {}
    ordered: List[str] = []
    for item in variants:
        if item and item not in seen:
            seen[item] = True
            ordered.append(item)
    return ordered


def ensure_default_country_code(digits: Optional[str]) -> Optional[str]:
    if not digits:
        return None
    sanitized = digits.lstrip("+")
    if sanitized.startswith("55"):
        return sanitized
    if len(sanitized) in {10, 11}:
        return f"55{sanitized}"
    return sanitized


def normalize_phone(value: Optional[str]) -> Optional[str]:
    variants = phone_variants(value)
    if not variants:
        return None
    enriched: List[str] = []
    for variant in variants:
        enriched.append(variant)
        with_country = ensure_default_country_code(variant)
        if with_country:
            enriched.append(with_country)
    filtered = [item for item in enriched if item]
    if not filtered:
        return None
    filtered = sorted(set(filtered), key=len, reverse=True)
    return filtered[0]


def find_conversation_by_phone(
    session: Session,
    organization_id: int,
    phone_number: Optional[str],
) -> Optional[Conversation]:
    normalized = normalize_phone(phone_number)
    candidates = phone_variants(normalized or phone_number)
    if not candidates:
        return None
    return session.exec(
        select(Conversation)
        .where(Conversation.organization_id == organization_id)
        .where(Conversation.debtor_phone.in_(candidates))
        .order_by(Conversation.updated_at.desc())
    ).first()


def ensure_conversation_for_phone(
    session: Session,
    organization_id: int,
    phone_number: str,
    *,
    display_name: Optional[str],
    owner_user_id: int,
) -> Conversation:
    conversation = find_conversation_by_phone(session, organization_id, phone_number)
    if conversation:
        if not conversation.owner_user_id:
            conversation.owner_user_id = owner_user_id
            session.add(conversation)
            session.commit()
            session.refresh(conversation)
        return conversation
    conversation = Conversation(
        organization_id=organization_id,
        debtor_name=display_name or phone_number or "Contato",
        debtor_phone=normalize_phone(phone_number) or phone_number,
        owner_user_id=owner_user_id,
    )
    session.add(conversation)
    session.commit()
    session.refresh(conversation)
    return conversation


PLACEHOLDER_PATTERN = re.compile(r"\{([^{}]+)\}")
SUPPORTED_SPREADSHEET_EXTENSIONS = {".csv", ".xlsx", ".xlsm", ".xls"}
MAX_UPLOAD_SIZE_BYTES = 10 * 1024 * 1024
MAX_CAMPAIGN_CONTACTS = 5000
CSV_DELIMITER_CANDIDATES = [",", ";", "\t", "|"]


def slugify_placeholder(value: Optional[str]) -> str:
    if not value:
        return ""
    normalized = re.sub(r"[^a-z0-9]+", "_", value.strip().lower())
    return normalized.strip("_")


def normalize_column_key(value: Optional[str]) -> str:
    if not value:
        return ""
    simplified = re.sub(r"\s+", " ", str(value)).strip().lower()
    return simplified


def _ensure_mapping_column(column: Optional[str]) -> str:
    if not column:
        raise HTTPException(status_code=400, detail="Selecione a coluna correspondente na planilha.")
    sanitized = column.strip()
    if not sanitized:
        raise HTTPException(status_code=400, detail="Selecione a coluna correspondente na planilha.")
    return sanitized


def parse_contacts_from_rows(
    rows: List[Dict[str, str]],
    mapping: BulkContactsMapping,
) -> List[Dict[str, Any]]:
    if not rows:
        raise HTTPException(status_code=400, detail="Nenhum contato valido foi encontrado no arquivo.")
    requested_phone_column = _ensure_mapping_column(mapping.phone)
    missing_columns: List[str] = []
    sample = rows[0]
    column_lookup: Dict[str, str] = {}
    for column_name in sample.keys():
        normalized = normalize_column_key(column_name)
        if normalized and normalized not in column_lookup:
            column_lookup[normalized] = column_name

    def resolve_column(column_label: Optional[str]) -> Optional[str]:
        if not column_label:
            return None
        normalized = normalize_column_key(column_label)
        return column_lookup.get(normalized)

    phone_column = resolve_column(requested_phone_column)
    if not phone_column:
        missing_columns.append(requested_phone_column)
    requested_name_column = mapping.name.strip() if mapping.name else None
    name_column = resolve_column(requested_name_column) if requested_name_column else None
    if requested_name_column and not name_column:
        missing_columns.append(requested_name_column)
    resolved_custom_fields: List[Tuple[BulkCustomFieldMapping, str]] = []
    for custom in mapping.custom_fields:
        actual = resolve_column(custom.column)
        if not actual:
            missing_columns.append(custom.column)
            continue
        resolved_custom_fields.append((custom, actual))
    if missing_columns:
        raise HTTPException(
            status_code=400,
            detail=f"Colunas nao encontradas na planilha: {', '.join(sorted(set(missing_columns)))}",
        )
    contacts: List[Dict[str, Any]] = []
    for row in rows:
        raw_phone = str(row.get(phone_column) or "").strip()
        normalized_phone = normalize_phone(raw_phone)
        if not normalized_phone:
            continue
        entry: Dict[str, Any] = {
            "phone": normalized_phone,
            "raw_phone": raw_phone,
            "name": str(row.get(name_column) or "").strip() if name_column else "",
            "fields": {},
        }
        for custom, column_name in resolved_custom_fields:
            key_slug = slugify_placeholder(custom.key or custom.column)
            if not key_slug:
                continue
            entry["fields"][key_slug] = str(row.get(column_name) or "").strip()
        contacts.append(entry)
    if not contacts:
        raise HTTPException(status_code=400, detail="Nenhum telefone valido foi encontrado no arquivo.")
    return contacts


def parse_csv_bytes(blob: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    text = blob.decode("utf-8-sig", errors="ignore")
    raw_lines = [line for line in text.splitlines() if line.strip()]
    delimiter = detect_csv_delimiter(raw_lines[0] if raw_lines else "")
    stream = io.StringIO(text)
    reader = csv.DictReader(stream, delimiter=delimiter)
    headers = reader.fieldnames
    if not headers:
        raise HTTPException(status_code=400, detail="Nao foi possivel identificar o cabecalho da planilha.")
    rows: List[Dict[str, str]] = []
    for row in reader:
        cleaned = {header: str(row.get(header) or "").strip() for header in headers}
        if any(cleaned.values()):
            rows.append(cleaned)
    return headers, rows


def detect_csv_delimiter(sample_line: str) -> str:
    if not sample_line:
        return ","
    best_delimiter = ","
    best_score = -1
    for delimiter in CSV_DELIMITER_CANDIDATES:
        score = sample_line.count(delimiter)
        if score > best_score:
            best_delimiter = delimiter
            best_score = score
    return best_delimiter or ","


def parse_xlsx_bytes(blob: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover - optional dependency
        raise HTTPException(status_code=500, detail="Suporte a arquivos XLSX nao disponivel.") from exc
    workbook = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise HTTPException(status_code=400, detail="Arquivo sem cabecalho ou vazio.")
    headers: List[str] = []
    seen: Dict[str, int] = {}
    for index, cell in enumerate(header_row):
        base_label = str(cell).strip() if cell is not None else f"Coluna {index + 1}"
        if not base_label:
            base_label = f"Coluna {index + 1}"
        candidate = base_label
        while candidate in seen:
            seen[candidate] += 1
            candidate = f"{base_label} ({seen[candidate]})"
        seen[candidate] = 1
        headers.append(candidate)
    rows: List[Dict[str, str]] = []
    for row_values in rows_iter:
        row_dict: Dict[str, str] = {}
        row_has_data = False
        for idx, header in enumerate(headers):
            cell_value = row_values[idx] if idx < len(row_values) else None
            if cell_value is None:
                value = ""
            else:
                value = str(cell_value).strip()
            if value:
                row_has_data = True
            row_dict[header] = value
        if row_has_data:
            rows.append(row_dict)
    return headers, rows


def extract_rows_from_file(data: bytes, filename: str) -> Tuple[List[str], List[Dict[str, str]]]:
    suffix = (Path(filename).suffix or "").lower()
    if suffix == ".csv":
        return parse_csv_bytes(data)
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return parse_xlsx_bytes(data)
    valid = ", ".join(sorted(SUPPORTED_SPREADSHEET_EXTENSIONS))
    raise HTTPException(
        status_code=400,
        detail=f"Formato de arquivo nao suportado ({suffix or 'sem extensao'}). Use um dos formatos: {valid}.",
    )


def build_template_context(contact: Dict[str, Any]) -> Dict[str, str]:
    context: Dict[str, str] = {}
    if contact.get("name"):
        context["nome"] = contact["name"]
    context["telefone"] = contact["phone"]
    for key, value in contact.get("fields", {}).items():
        if key:
            context[key] = value
    return context


def render_message_template(template: str, context: Dict[str, str]) -> str:
    def replace(match: re.Match[str]) -> str:
        key = slugify_placeholder(match.group(1))
        if not key:
            return ""
        return context.get(key, "")

    return PLACEHOLDER_PATTERN.sub(replace, template)


def visible_tags_for_user(tags: Iterable[Tag], user: User) -> List[TagRead]:
    visible: List[TagRead] = []
    for tag in tags or []:
        if tag.owner_user_id not in (None, user.id):
            continue
        visible.append(TagRead.model_validate(tag))
    return visible


def extract_int(payload: Dict[str, Any], *candidates: str) -> Optional[int]:
    for key in candidates:
        if key in payload and payload[key] is not None:
            try:
                return int(payload[key])
            except (TypeError, ValueError):
                continue
    return None


def map_remote_campaign_status(raw_status: Optional[str]) -> Optional[str]:
    if not raw_status:
        return None
    normalized = str(raw_status).strip().lower()
    if normalized in {"running", "processing", "active"}:
        return "running"
    if normalized in {"queued", "pending", "scheduled"}:
        return "scheduled"
    if normalized in {"finished", "done", "completed", "archived"}:
        return "done"
    if normalized in {"stopped", "paused"}:
        return "paused"
    return normalized or None


def maybe_sync_bulk_campaigns(
    db_session: Session,
    current_user: User,
    campaigns: List[BulkCampaign],
) -> None:
    if not campaigns:
        return
    if not any(c.folder_id for c in campaigns):
        return
    integration_session = ensure_session_exists(
        db_session, current_user.organization_id, current_user.id
    )
    client = get_gateway_client(integration_session)
    if not isinstance(client, UazapiClient) or not client.is_configured:
        return
    try:
        remote_entries = client.list_mass_campaigns()
    except Exception as exc:  # pragma: no cover - integracao externa
        logger.warning("Nao foi possivel sincronizar status da Uazapi: %s", exc)
        return
    remote_index: Dict[str, Dict[str, Any]] = {}
    for entry in remote_entries or []:
        remote_id = entry.get("folder_id") or entry.get("folderId") or entry.get("id") or entry.get("folder")
        if remote_id:
            remote_index[str(remote_id)] = entry
    updated = False
    now = datetime.utcnow()
    for campaign in campaigns:
        if not campaign.folder_id:
            continue
        remote = remote_index.get(campaign.folder_id)
        if not remote:
            continue
        updates: Dict[str, Any] = {}
        mapped_status = map_remote_campaign_status(remote.get("status") or remote.get("state"))
        if mapped_status and mapped_status != campaign.status:
            updates["status"] = mapped_status
            if mapped_status == "running" and not campaign.started_at:
                updates["started_at"] = now
            if mapped_status == "done" and not campaign.finished_at:
                updates["finished_at"] = now
        sent_count = extract_int(remote, "sent", "delivered", "processed")
        if sent_count is not None and sent_count != campaign.sent_count:
            updates["sent_count"] = sent_count
        failed_count = extract_int(remote, "failed", "errors")
        if failed_count is not None and failed_count != campaign.failed_count:
            updates["failed_count"] = failed_count
        total = extract_int(remote, "total", "count", "queued")
        if total is not None and total != campaign.contacts_total:
            updates["contacts_total"] = total
        if updates:
            for key, value in updates.items():
                setattr(campaign, key, value)
            campaign.last_synced_at = now
            campaign.updated_at = now
            db_session.add(campaign)
            updated = True
    if updated:
        db_session.commit()


def extract_evolution_text(message_payload: Dict[str, Any]) -> Optional[str]:
    if not message_payload:
        return None
    if "conversation" in message_payload:
        return message_payload.get("conversation")
    extended = message_payload.get("extendedTextMessage")
    if isinstance(extended, dict):
        text = extended.get("text") or extended.get("caption")
        if text:
            return text
    for field in ("imageMessage", "videoMessage", "documentMessage"):
        media = message_payload.get(field)
        if isinstance(media, dict):
            caption = media.get("caption") or media.get("title")
            if caption:
                return caption
    buttons = message_payload.get("buttonsResponseMessage")
    if isinstance(buttons, dict):
        text = buttons.get("selectedDisplayText") or buttons.get("selectedButtonId")
        if text:
            return text
    template_reply = message_payload.get("templateButtonReplyMessage")
    if isinstance(template_reply, dict):
        text = template_reply.get("selectedDisplayText") or template_reply.get("selectedId")
        if text:
            return text
    return None


def extract_evolution_audio(message_payload: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if not message_payload:
        return None
    for field in ("audioMessage", "pttMessage", "voiceMessage"):
        media = message_payload.get(field)
        if isinstance(media, dict):
            return media
    document = message_payload.get("documentMessage")
    if isinstance(document, dict):
        mimetype = (document.get("mimetype") or "").lower()
        if mimetype.startswith("audio/"):
            return document
    return None


def extract_evolution_document(message_payload: Dict[str, Any]) -> Optional[Tuple[Dict[str, Any], MessageType]]:
    if not message_payload:
        return None
    document = message_payload.get("documentMessage")
    if isinstance(document, dict):
        return document, MessageType.document
    image = message_payload.get("imageMessage")
    if isinstance(image, dict):
        return image, MessageType.image
    video = message_payload.get("videoMessage")
    if isinstance(video, dict):
        return video, MessageType.document
    return None


def extract_uazapi_text(message_payload: Dict[str, Any]) -> Optional[str]:
    if not message_payload:
        return None
    for key in ("text", "body", "message", "content", "caption"):
        value = message_payload.get(key)
        if isinstance(value, str) and value.strip():
            return value
    nested = message_payload.get("content")
    if isinstance(nested, dict):
        for key in ("text", "caption", "body"):
            value = nested.get(key)
            if isinstance(value, str) and value.strip():
                return value
    return None


def clean_message_identifier(value: Optional[Any]) -> Optional[str]:
    if isinstance(value, str):
        cleaned = value.strip()
        if cleaned:
            return cleaned
    return None


def extract_uazapi_reply_reference(message_payload: Dict[str, Any]) -> Optional[str]:
    if not isinstance(message_payload, dict):
        return None
    for key in ("quoted", "quotedMessageId", "replyid", "replyId"):
        candidate = clean_message_identifier(message_payload.get(key))
        if candidate:
            return candidate
    candidate_contexts: List[Dict[str, Any]] = []
    direct_context = message_payload.get("contextInfo")
    if isinstance(direct_context, dict):
        candidate_contexts.append(direct_context)
    nested_content = message_payload.get("content")
    if isinstance(nested_content, dict):
        ctx = nested_content.get("contextInfo")
        if isinstance(ctx, dict):
            candidate_contexts.append(ctx)
    generic_context = message_payload.get("context")
    if isinstance(generic_context, dict):
        candidate_contexts.append(generic_context)
    for ctx in candidate_contexts:
        for key in ("stanzaId", "stanzaID", "quotedMessageId", "quotedStanzaId"):
            candidate = clean_message_identifier(ctx.get(key))
            if candidate:
                return candidate
        quoted_payload = ctx.get("quotedMessage")
        if isinstance(quoted_payload, dict):
            nested = extract_uazapi_reply_reference(quoted_payload)
            if nested:
                return nested
    return None


def extract_evolution_reply_reference(raw_message: Dict[str, Any]) -> Optional[str]:
    if not isinstance(raw_message, dict):
        return None

    def walker(payload: Dict[str, Any]) -> Optional[str]:
        if not isinstance(payload, dict):
            return None
        context = payload.get("contextInfo")
        if isinstance(context, dict):
            for key in ("stanzaId", "stanzaID", "quotedMessageId", "quotedStanzaId"):
                candidate = clean_message_identifier(context.get(key))
                if candidate:
                    return candidate
            quoted_message = context.get("quotedMessage")
            if isinstance(quoted_message, dict):
                nested = walker(quoted_message)
                if nested:
                    return nested
        for value in payload.values():
            if isinstance(value, dict):
                nested = walker(value)
                if nested:
                    return nested
        return None

    return walker(raw_message.get("message") or raw_message)


IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".heic"}


def infer_uploaded_media_type(filename: Optional[str], content_type: Optional[str]) -> Tuple[MessageType, str]:
    lowered_type = (content_type or "").lower()
    extension = (Path(filename or "").suffix or "").lower()
    if lowered_type.startswith("image/") or extension in IMAGE_EXTENSIONS:
        mime = lowered_type or "image/jpeg"
        return MessageType.image, mime
    if lowered_type in {"application/pdf", "application/x-pdf"} or extension == ".pdf":
        return MessageType.document, "application/pdf"
    raise HTTPException(status_code=400, detail="Apenas imagens e PDFs sao suportados.")


def detect_uazapi_message_type(message_payload: Dict[str, Any]) -> MessageType:
    candidates = [
        message_payload.get("messageType"),
        message_payload.get("type"),
        message_payload.get("mediaType"),
        message_payload.get("mimetype"),
        message_payload.get("mimeType"),
    ]
    nested = message_payload.get("content")
    if isinstance(nested, dict):
        candidates.extend(
            nested.get(key)
            for key in ("type", "messageType", "mimetype", "mimeType", "mediaType")
        )
    normalized = [str(value).lower() for value in candidates if isinstance(value, str)]
    for value in normalized:
        if any(token in value for token in ("audio", "voice", "ptt", "push_to_talk", "push-to-talk")):
            return MessageType.audio
        if any(token in value for token in ("image", "photo", "picture")):
            return MessageType.image
        if any(token in value for token in ("document", "application/", "pdf")):
            return MessageType.document
    return MessageType.text


def determine_incoming_filename(
    original_name: Optional[str],
    message_type: MessageType,
    content_type: Optional[str],
) -> str:
    if original_name:
        original_name = original_name.strip()
        if original_name:
            if Path(original_name).suffix:
                return original_name
            base_name = Path(original_name).stem or original_name
        else:
            base_name = f"{message_type.value}-message"
    else:
        base_name = f"{message_type.value}-message"
    mime = (content_type or "").split(";", 1)[0].strip()
    extension = guess_extension(mime) or ""
    if not extension:
        if message_type == MessageType.audio:
            extension = ".ogg"
        elif message_type == MessageType.image:
            extension = ".jpg"
        elif message_type == MessageType.document:
            extension = ".pdf"
        else:
            extension = ".bin"
    return f"{base_name}{extension}"


NAME_BLACKLIST = {"unknown", "desconhecido"}


def normalize_display_name(value: Optional[str]) -> Optional[str]:
    if not isinstance(value, str):
        return None
    trimmed = value.strip()
    if not trimmed:
        return None
    lowered = trimmed.lower()
    if lowered in NAME_BLACKLIST:
        return None
    return trimmed[:120]


def resolve_uazapi_media_url(file_url: str, integration_session: DeviceSession) -> str:
    if not file_url:
        return ""
    file_url = file_url.strip()
    if file_url.startswith("http://") or file_url.startswith("https://") or file_url.startswith("data:"):
        return file_url
    base_url = (integration_session.integration_base_url or "").rstrip("/")
    if not base_url:
        return file_url
    if not file_url.startswith("/"):
        file_url = f"/{file_url}"
    return f"{base_url}{file_url}"


def download_uazapi_media(
    message_payload: Dict[str, Any],
    integration_session: DeviceSession,
    *,
    expected_type: Optional[MessageType] = None,
) -> Optional[Tuple[bytes, Optional[str], Optional[str]]]:
    file_url = (
        message_payload.get("fileURL")
        or message_payload.get("fileUrl")
        or message_payload.get("url")
        or message_payload.get("mediaUrl")
    )
    if not file_url:
        content = message_payload.get("content")
        if isinstance(content, dict):
            file_url = (
                content.get("fileURL")
                or content.get("fileUrl")
                or content.get("url")
                or content.get("mediaUrl")
            )
    filename = message_payload.get("docName") or message_payload.get("fileName")
    message_id = message_payload.get("messageid") or message_payload.get("id")
    resolved = resolve_uazapi_media_url(file_url, integration_session) if file_url else None

    attempts: List[Tuple[str, callable]] = []
    if resolved and resolved.startswith("data:"):
        attempts.append(("data-uri", lambda: _decode_inline_media(resolved)))
    if message_id:
        attempts.append(
            (
                "uazapi-download",
                lambda: _download_uazapi_message_media(message_id, integration_session, filename),
            )
        )
    if resolved:
        attempts.append(
            (
                "http-download",
                lambda: _download_http_media(resolved, integration_session),
            )
        )

    for label, loader in attempts:
        try:
            result = loader()
        except Exception as exc:  # defensive
            logger.warning("Falha ao tentar %s para mensagem %s: %s", label, message_id, exc)
            continue
        if not result:
            continue
        data, content_type, detected_name = result
        if _should_discard_download(data, content_type, expected_type):
            logger.warning(
                "Descartando midia baixada via %s para mensagem %s por conteudo inesperado (%s)",
                label,
                message_id,
                content_type,
            )
            continue
        return data, content_type, detected_name or filename
    return None


def _decode_inline_media(value: str) -> Optional[Tuple[bytes, Optional[str]]]:
    header, _, encoded = value.partition(",")
    if not encoded:
        return None
    content_type = None
    if header.startswith("data:"):
        content_type = header[5:].split(";", 1)[0] or None
    try:
        return base64.b64decode(encoded), content_type
    except Exception:
        return None


def _download_http_media(
    resolved: str,
    integration_session: DeviceSession,
) -> Optional[Tuple[bytes, Optional[str], Optional[str]]]:
    headers: Dict[str, str] = {
        "User-Agent": "chatdevalor-bot/1.0",
        "ngrok-skip-browser-warning": "true",
    }
    if integration_session.integration_token:
        headers["token"] = integration_session.integration_token
    try:
        response = requests.get(resolved, headers=headers, timeout=60)
        response.raise_for_status()
    except requests.RequestException as exc:
        logger.warning("Erro ao baixar media da Uazapi: %s", exc)
        return None
    content_type = response.headers.get("Content-Type")
    if _looks_like_ngrok_warning(response.content, content_type):
        logger.warning("Download bloqueado pelo aviso do Ngrok para %s", resolved)
        return None
    filename: Optional[str] = None
    try:
        parsed = urlparse(resolved)
        filename = Path(parsed.path or "").name or None
    except Exception:
        filename = None
    return response.content, content_type, filename


def _looks_like_ngrok_warning(payload: bytes, content_type: Optional[str]) -> bool:
    if not payload:
        return False
    try:
        snippet = payload[:2048].decode("utf-8", errors="ignore").lower()
    except Exception:
        return False
    if "ngrok.com" not in snippet:
        return False
    return "err_ngrok" in snippet or "you are about to visit" in snippet


def _should_discard_download(
    payload: Optional[bytes],
    content_type: Optional[str],
    expected_type: Optional[MessageType],
) -> bool:
    if not payload:
        return True
    if _looks_like_ngrok_warning(payload, content_type):
        return True
    normalized = (content_type or "").split(";", 1)[0].strip().lower()
    if expected_type == MessageType.document and normalized.startswith("text/html"):
        return True
    return False


def _download_uazapi_message_media(
    message_id: str,
    integration_session: DeviceSession,
    fallback_filename: Optional[str] = None,
) -> Optional[Tuple[bytes, Optional[str], Optional[str]]]:
    base_url = (integration_session.integration_base_url or "").rstrip("/")
    token = integration_session.integration_token
    if not base_url or not token:
        return None
    url = f"{base_url}/message/download"
    payload: Dict[str, Any] = {
        "id": message_id,
        "messageId": message_id,
        "return_base64": True,
        "return_link": True,
        "generate_mp3": False,
    }
    if integration_session.integration_instance_id:
        payload["session"] = integration_session.integration_instance_id
        payload["sessionId"] = integration_session.integration_instance_id
        payload["session_id"] = integration_session.integration_instance_id
        payload["instance"] = integration_session.integration_instance_id
        payload["instanceId"] = integration_session.integration_instance_id
    headers = {
        "accept": "application/json",
        "Content-Type": "application/json",
        "token": token,
    }
    try:
        response = requests.post(url, headers=headers, json=payload, timeout=60)
        response.raise_for_status()
    except requests.RequestException as exc:
        logger.warning("Erro ao usar /message/download Uazapi: %s", exc)
        return None
    try:
        body = response.json()
    except ValueError:
        logger.warning("Resposta download Uazapi nao era JSON")
        return None

    base64_fields = [
        body.get("base64Data"),
        body.get("base64"),
        body.get("fileBase64"),
        body.get("data"),
    ]
    for field in base64_fields:
        if not isinstance(field, str) or not field.strip():
            continue
        decoded = _decode_inline_media(field.strip()) or (_decode_inline_media(f"data:;base64,{field.strip()}"))
        if decoded:
            data, inferred_type = decoded
            content_type = body.get("mimetype") or body.get("mimeType") or inferred_type
            filename = body.get("fileName") or fallback_filename
            return data, content_type, filename

    link = body.get("fileURL") or body.get("fileUrl") or body.get("url")
    if isinstance(link, str) and link.strip():
        resolved = resolve_uazapi_media_url(link.strip(), integration_session)
        return _download_http_media(resolved, integration_session)
    return None


def build_public_media_url(key: Optional[str], request: Optional[Request]) -> Optional[str]:
    if not key:
        return None
    candidate_url: Optional[str] = None
    try:
        candidate_url = media_storage.build_url(key)
    except Exception:
        candidate_url = None
    if candidate_url:
        if candidate_url.startswith("http://") or candidate_url.startswith("https://"):
            return candidate_url
        base = str(request.base_url).rstrip("/") if request else ""
        if candidate_url.startswith("/"):
            return f"{base}{candidate_url}"
        return f"{base}/{candidate_url}"
    if media_storage.is_local() and request:
        try:
            return str(request.url_for("media", path=key))
        except Exception:
            return None
    return None


def build_media_prefix(conversation: Conversation) -> str:
    org = conversation.organization_id or 0
    conv_id = conversation.id or "pending"
    return f"org_{org}/conv_{conv_id}"


def _extract_avatar_blob(
    image_reference: Optional[str],
    integration_session: DeviceSession,
) -> Optional[Tuple[bytes, Optional[str]]]:
    if not image_reference:
        return None
    value = image_reference.strip()
    if not value:
        return None
    if value.startswith("data:"):
        inline = _decode_inline_media(value)
        if inline:
            return inline
        return None
    if value.startswith("http://") or value.startswith("https://"):
        resolved = resolve_uazapi_media_url(value, integration_session)
        downloaded = _download_http_media(resolved, integration_session)
        if downloaded:
            data, content_type, _ = downloaded
            return data, content_type
        return None
    try:
        decoded = base64.b64decode(value, validate=False)
        if decoded:
            return decoded, "image/jpeg"
    except Exception:
        return None
    return None


def fetch_and_store_conversation_avatar(
    session: Session,
    conversation: Conversation,
    integration_session: DeviceSession,
    request: Request,
    *,
    preview: bool = True,
) -> Optional[str]:
    provider = (integration_session.provider or "").lower()
    if provider != "uazapi":
        return None
    client = get_gateway_client(integration_session)
    if not isinstance(client, UazapiClient):
        return None
    try:
        details = client.fetch_chat_details(conversation.debtor_phone, preview=preview)
    except requests.HTTPError as exc:  # pragma: no cover - rede externa
        body = ""
        if exc.response is not None:
            try:
                body = exc.response.text[:500]
            except Exception:
                body = ""
        logger.error("Uazapi chat/details falhou: %s body=%s", exc, body)
        raise HTTPException(status_code=502, detail="Falha ao consultar avatar na Uazapi.") from exc
    image_value = None
    if preview:
        image_value = details.get("imagePreview") or details.get("image")
    else:
        image_value = details.get("image") or details.get("imagePreview")
    blob = _extract_avatar_blob(image_value, integration_session)
    avatar_entry = session.get(ConversationAvatar, conversation.id)
    if not avatar_entry:
        avatar_entry = ConversationAvatar(conversation_id=conversation.id)
    if not blob:
        avatar_entry.media_path = None
        avatar_entry.media_content_type = None
        avatar_entry.updated_at = datetime.utcnow()
        session.add(avatar_entry)
        session.commit()
        return None
    data, content_type = blob
    mime = (content_type or "").split(";", 1)[0] or "image/jpeg"
    extension = guess_extension(mime) or ".jpg"
    stored_media = media_storage.save_bytes(
        data,
        filename=f"avatar-{conversation.id}{extension}",
        content_type=content_type or "image/jpeg",
        prefix=build_media_prefix(conversation),
    )
    avatar_entry.media_path = stored_media.key
    avatar_entry.media_content_type = stored_media.content_type
    avatar_entry.updated_at = datetime.utcnow()
    session.add(avatar_entry)
    session.commit()
    return build_public_media_url(stored_media.key, request)


def append_webhook_dump(provider: str, payload: Any) -> None:
    try:
        WEBHOOK_LOG_DIR.mkdir(parents=True, exist_ok=True)
        target = WEBHOOK_LOG_DIR / f"{provider}-webhook.jsonl"
        with target.open("a", encoding="utf-8") as handle:
            record = {
                "ts": datetime.utcnow().isoformat(),
                "payload": payload,
            }
            handle.write(json.dumps(record, ensure_ascii=False))
            handle.write("\n")
    except Exception as exc:  # pragma: no cover - defensive
        logger.warning("Nao foi possivel registrar webhook %s: %s", provider, exc)



def serialize_message_entity(
    message: Message,
    deletion_entry: Optional[MessageDeletion] = None,
    *,
    my_reaction: Optional[str] = None,
    reaction_counts: Optional[Dict[str, int]] = None,
) -> MessageRead:
    data = MessageRead.model_validate(message)
    if message.media_path:
        data.media_url = media_storage.build_url(message.media_path)
    data.media_path = None
    if deletion_entry and deletion_entry.deleted_for_all:
        data.is_deleted_for_all = True
        data.deleted_for_all_at = deletion_entry.deleted_at
    if my_reaction is not None:
        data.my_reaction = my_reaction
    if reaction_counts is not None:
        data.reaction_counts = reaction_counts
    return data


def extract_provider_message_id(payload: Any) -> Optional[str]:
    """
    Attempt to extract a WhatsApp message id from a provider payload.
    Providers usually return keys such as `messageId`, `messageid`, `id` or nested `key`.
    """

    def is_candidate(value: Any) -> bool:
        if not isinstance(value, str):
            return False
        # WhatsApp message ids are usually alphanumeric/hex strings with at least 8 chars
        return len(value.strip()) >= 8

    def walker(obj: Any) -> Optional[str]:
        if isinstance(obj, dict):
            preferred_keys = (
                "messageid",
                "messageId",
                "id",
                "ID",
                "key",
                "response",
                "data",
            )
            for key in preferred_keys:
                if key in obj:
                    value = obj[key]
                    if is_candidate(value):
                        return value
                    nested = walker(value)
                    if nested:
                        return nested
            for value in obj.values():
                nested = walker(value)
                if nested:
                    return nested
        elif isinstance(obj, (list, tuple)):
            for item in obj:
                nested = walker(item)
                if nested:
                    return nested
        return None

    return walker(payload)


def serialize_session(device_session: DeviceSession) -> SessionRead:
    session_read = SessionRead.model_validate(device_session)
    session_read.has_credentials = session_has_credentials(device_session)
    session_read.webhook_id = device_session.webhook_id
    return session_read


def find_session_by_instance(session: Session, instance_identifier: Optional[str]) -> Optional[DeviceSession]:
    if not instance_identifier:
        return None
    return session.exec(
        select(DeviceSession)
        .where(DeviceSession.integration_instance_id == instance_identifier)
        .order_by(DeviceSession.created_at.desc())
    ).first()


def get_gateway_client(integration_session: Optional[DeviceSession]):
    provider = (integration_session.provider if integration_session else "uazapi") or "uazapi"
    provider = provider.lower()
    if provider == "uazapi" and integration_session:
        uazapi_client = UazapiClient(
            base_url=integration_session.integration_base_url,
            instance_id=integration_session.integration_instance_id,
            token=integration_session.integration_token,
        )
        if uazapi_client.is_configured:
            return uazapi_client

    if provider == "wppconnect" and integration_session:
        candidate_wpp = WPPConnectClient(
            base_url=integration_session.integration_base_url,
            session_name=integration_session.integration_instance_id,
            token=integration_session.integration_token,
        )
        if candidate_wpp.is_configured:
            return candidate_wpp

    if integration_session:
        candidate = EvolutionClient.from_values(
            base_url=integration_session.integration_base_url,
            instance_id=integration_session.integration_instance_id,
            token=integration_session.integration_token,
            admin_token=evolution_client.admin_token,
        )
        if candidate.is_configured:
            return candidate
    return evolution_client


def resolve_uazapi_session_for_conversation(
    session: Session,
    *,
    conversation: Conversation,
    current_user: User,
) -> Optional[DeviceSession]:
    """Best-effort resolver for the Uazapi session when message.via_session_id is missing.

    Some older records or webhook-ingested messages may not have `via_session_id` set.
    For Uazapi-only operations (delete/edit/react), we try to locate the most recent
    configured Uazapi session for the conversation owner (fallback to current user,
    then any configured session in the organization).
    """

    def query_for_owner(owner_id: Optional[int]) -> Optional[DeviceSession]:
        if not owner_id:
            return None
        candidate = session.exec(
            select(DeviceSession)
            .where(DeviceSession.organization_id == conversation.organization_id)
            .where(DeviceSession.provider == "uazapi")
            .where(DeviceSession.owner_user_id == owner_id)
            .where(DeviceSession.integration_base_url.is_not(None))
            .where(DeviceSession.integration_instance_id.is_not(None))
            .where(DeviceSession.integration_token.is_not(None))
            .order_by(DeviceSession.created_at.desc())
        ).first()
        return candidate if (candidate and session_has_credentials(candidate)) else None

    candidate = query_for_owner(conversation.owner_user_id)
    if candidate:
        return candidate
    candidate = query_for_owner(current_user.id)
    if candidate:
        return candidate

    any_candidate = session.exec(
        select(DeviceSession)
        .where(DeviceSession.organization_id == conversation.organization_id)
        .where(DeviceSession.provider == "uazapi")
        .where(DeviceSession.integration_base_url.is_not(None))
        .where(DeviceSession.integration_instance_id.is_not(None))
        .where(DeviceSession.integration_token.is_not(None))
        .order_by(DeviceSession.created_at.desc())
    ).first()
    return any_candidate if (any_candidate and session_has_credentials(any_candidate)) else None


def extract_provider_message_id(payload: Any) -> Optional[str]:
    """Busca recursivamente um identificador retornado pelo gateway."""

    def is_candidate(value: Any) -> bool:
        return isinstance(value, str) and len(value.strip()) >= 8

    def walker(obj: Any) -> Optional[str]:
        if isinstance(obj, dict):
            preferred = (
                "messageid",
                "messageId",
                "id",
                "ID",
                "key",
                "response",
                "data",
            )
            for key in preferred:
                if key in obj:
                    value = obj[key]
                    if is_candidate(value):
                        return value
                    nested = walker(value)
                    if nested:
                        return nested
            for value in obj.values():
                nested = walker(value)
                if nested:
                    return nested
        elif isinstance(obj, (list, tuple)):
            for item in obj:
                nested = walker(item)
                if nested:
                    return nested
        return None

    return walker(payload)


def download_media_from_evolution(
    media_payload: Dict[str, Any],
    integration_session: DeviceSession,
) -> Optional[tuple[bytes, str, str]]:
    client = get_gateway_client(integration_session)
    if not client.is_configured:
        return None
    debug_info: Dict[str, Any] = {
        "payload_keys": list(media_payload.keys()),
        "instance": integration_session.integration_instance_id,
        "mimetype": media_payload.get("mimetype"),
        "fileName": media_payload.get("fileName"),
    }
    media_url: Optional[str] = None
    for key in (
        "directPlayUrl",
        "mediaUrl",
        "url",
        "downloadUrl",
        "mediaDirectUrl",
    ):
        raw = media_payload.get(key)
        if isinstance(raw, str) and raw:
            media_url = raw
            debug_info["media_url_key"] = key
            break
    if not media_url:
        encoded = media_payload.get("fileBase64") or media_payload.get("base64")
        if isinstance(encoded, str) and encoded:
            try:
                data = base64.b64decode(
                    encoded.split(",", 1)[-1], validate=False
                )
                mime = media_payload.get("mimetype") or "application/octet-stream"
                filename = media_payload.get("fileName") or "media.bin"
                debug_info["source"] = "inline_base64"
                debug_info["decoded_bytes"] = len(data)
                _write_media_debug(debug_info)
                return data, mime, filename
            except Exception:
                debug_info["error"] = "base64_decode_failed"
                _write_media_debug(debug_info)
                return None
        debug_info["error"] = "no_media_reference"
        _write_media_debug(debug_info)
        return None
    if media_url.startswith("data:"):
        try:
            header, encoded = media_url.split(",", 1)
            mime_part = header.split(";")[0]
            mime = mime_part[5:] if mime_part.startswith("data:") else "application/octet-stream"
            data = base64.b64decode(encoded, validate=False)
            filename = media_payload.get("fileName") or "media.bin"
            debug_info["source"] = "data_uri"
            debug_info["decoded_bytes"] = len(data)
            _write_media_debug(debug_info)
            return data, mime, filename
        except Exception:
            return None
    data, headers = client.download_media(media_url, return_headers=True)
    debug_info.update(
        {
            "source": "download",
            "media_url": media_url,
            "response_headers": dict(headers or {}),
            "bytes": len(data),
            "prefix": data[:32].hex(),
        }
    )
    _write_media_debug(debug_info)
    mime = media_payload.get("mimetype") or (headers or {}).get("Content-Type") or "audio/ogg"
    filename = media_payload.get("fileName") or "audio-message.ogg"
    return data, mime, filename


def _write_media_debug(info: Dict[str, Any]) -> None:
    try:
        MEDIA_DEBUG_FILE.parent.mkdir(parents=True, exist_ok=True)
        MEDIA_DEBUG_FILE.write_text(json.dumps(info, ensure_ascii=False, indent=2))
    except Exception:
        pass


def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)


def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)


def extract_bearer_token(authorization: Optional[str]) -> str:
    if not authorization:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token ausente")
    scheme, _, token = authorization.partition(" ")
    if scheme.lower() != "bearer" or not token:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token invalido")
    return token.strip()


def create_session_token(user: User, session: Session) -> SessionToken:
    token = SessionToken(token=str(uuid4()), user_id=user.id)
    session.add(token)
    session.commit()
    session.refresh(token)
    return token


def get_current_user(
    authorization: Optional[str] = Header(default=None),
    session: Session = Depends(get_session),
) -> User:
    token_value = extract_bearer_token(authorization)
    session_token = session.exec(select(SessionToken).where(SessionToken.token == token_value)).first()
    if not session_token:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Sessao invalida")
    user = session.get(User, session_token.user_id)
    if not user or not user.is_active:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Usuario inativo")
    return user


@app.get("/", response_class=HTMLResponse)
def serve_frontend() -> str:
    index_file = frontend_dir / "index.html"
    if not index_file.exists():
        raise HTTPException(status_code=404, detail="Interface nao encontrada")
    return index_file.read_text(encoding="utf-8")


@app.get("/colaboradores.html", response_class=HTMLResponse)
def serve_collaboradores() -> str:
    collab_file = frontend_dir / "colaboradores.html"
    if not collab_file.exists():
        raise HTTPException(status_code=404, detail="Área de colaboradores indisponível")
    return collab_file.read_text(encoding="utf-8")


if frontend_dir.exists():
    app.mount("/static", StaticFiles(directory=frontend_dir), name="static")


@app.get("/api/health")
def healthcheck() -> dict:
    return {"status": "ok", "timestamp": datetime.utcnow().isoformat()}


@app.get("/api/events")
async def stream_events(
    request: Request,
    current_user: User = Depends(get_current_user),
) -> StreamingResponse:
    org_id = current_user.organization_id
    if not org_id:
        raise HTTPException(status_code=400, detail="Organizacao invalida")

    queue = await _sse_register(org_id)

    async def event_stream() -> Iterable[str]:
        # Primeiro evento para o cliente saber que conectou.
        yield "event: ready\ndata: {}\n\n"
        try:
            while True:
                if await request.is_disconnected():
                    break
                try:
                    event_name, payload = await asyncio.wait_for(
                        queue.get(), timeout=SSE_PING_INTERVAL_SECONDS
                    )
                except asyncio.TimeoutError:
                    yield "event: ping\ndata: {}\n\n"
                    continue
                try:
                    data = json.dumps(payload or {}, ensure_ascii=False)
                except Exception:
                    data = "{}"
                yield f"event: {event_name}\ndata: {data}\n\n"
        finally:
            await _sse_unregister(org_id, queue)

    headers = {
        "Cache-Control": "no-cache",
        "Connection": "keep-alive",
        # Evita buffering em alguns proxies.
        "X-Accel-Buffering": "no",
    }
    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers=headers,
    )


@app.get("/api/conversations", response_model=List[ConversationSummary])
def list_conversations(
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    scope: str = Query("mine", regex="^(mine|all)$"),
) -> List[ConversationSummary]:
    query = select(Conversation).where(
        Conversation.organization_id == current_user.organization_id
    )
    if not current_user.is_admin or scope != "all":
        query = query.where(Conversation.owner_user_id == current_user.id)
    conversations = session.exec(
        query.order_by(Conversation.updated_at.desc())
    ).all()
    conversation_ids = [conversation.id for conversation in conversations]
    pinned_ids: set[int] = set()
    pinned_at_map: dict[int, datetime] = {}
    if conversation_ids:
        pinned_rows = session.exec(
            select(ConversationPin).where(
                ConversationPin.user_id == current_user.id,
                ConversationPin.conversation_id.in_(conversation_ids),
            )
        ).all()
        pinned_at_map = {
            entry.conversation_id: entry.created_at
            for entry in pinned_rows
            if entry.conversation_id is not None
        }
        pinned_ids = set(pinned_at_map.keys())
        conversations.sort(
            key=lambda conversation: (
                conversation.id in pinned_ids,
                pinned_at_map.get(conversation.id, datetime.min),
                conversation.updated_at,
            ),
            reverse=True,
        )
    unread_counts: Dict[int, int] = {}
    if conversation_ids:
        unread_rows = session.exec(
            select(Message.conversation_id, func.count(Message.id))
            .where(Message.conversation_id.in_(conversation_ids))
            .where(Message.direction == MessageDirection.debtor)
            .where(Message.is_read == False)  # noqa: E712
            .group_by(Message.conversation_id)
        ).all()
        unread_counts = {conversation_id: count for conversation_id, count in unread_rows}
    avatar_rows: List[ConversationAvatar] = []
    if conversation_ids:
        avatar_rows = session.exec(
            select(ConversationAvatar).where(ConversationAvatar.conversation_id.in_(conversation_ids))
        ).all()
    avatar_map: Dict[int, ConversationAvatar] = {
        entry.conversation_id: entry for entry in avatar_rows if entry.conversation_id
    }
    summaries: List[ConversationSummary] = []
    for conversation in conversations:
        tag_reads = visible_tags_for_user(conversation.tags, current_user)
        last_message = session.exec(
            select(Message)
            .where(Message.conversation_id == conversation.id)
            .order_by(Message.timestamp.desc())
        ).first()
        avatar_entry = avatar_map.get(conversation.id)
        avatar_url: Optional[str] = None
        avatar_updated_at = avatar_entry.updated_at if avatar_entry else None
        if avatar_entry and avatar_entry.media_path:
            avatar_url = build_public_media_url(avatar_entry.media_path, request)
        summaries.append(
            ConversationSummary(
                id=conversation.id,
                debtor_name=conversation.debtor_name,
                debtor_phone=conversation.debtor_phone,
                last_message_at=last_message.timestamp if last_message else None,
                last_message_preview=(
                    (last_message.content or "")[:default_message_preview_length]
                    if last_message
                    else None
                ),
                owner_user_id=conversation.owner_user_id,
                tags=tag_reads,
                unread_count=unread_counts.get(conversation.id, 0),
                is_pinned=conversation.id in pinned_ids,
                avatar_url=avatar_url,
                avatar_updated_at=avatar_updated_at,
            )
        )
    return summaries


@app.post("/api/conversations", response_model=ConversationRead)
def create_conversation(
    payload: ConversationCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ConversationRead:
    data = payload.model_dump()
    normalized_phone = normalize_phone(data.get("debtor_phone"))
    if normalized_phone:
        data["debtor_phone"] = normalized_phone
    conversation = Conversation(
        **data,
        organization_id=current_user.organization_id,
        owner_user_id=current_user.id,
    )
    session.add(conversation)
    session.commit()
    session.refresh(conversation)
    conversation_read = ConversationRead.model_validate(conversation)
    conversation_read.tags = visible_tags_for_user(conversation.tags, current_user)
    return conversation_read


@app.get("/api/conversations/{conversation_id}", response_model=ConversationRead)
def get_conversation(
    request: Request,
    conversation_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ConversationRead:
    conversation = session.get(Conversation, conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa nao encontrada")
    ensure_conversation_access(conversation, current_user)
    conversation_read = ConversationRead.model_validate(conversation)
    conversation_read.tags = visible_tags_for_user(conversation.tags, current_user)
    avatar_entry = session.get(ConversationAvatar, conversation_id)
    avatar_updated_at = avatar_entry.updated_at if avatar_entry else None
    avatar_url: Optional[str] = None
    if avatar_entry and avatar_entry.media_path:
        avatar_url = build_public_media_url(avatar_entry.media_path, request)
    conversation_read.avatar_url = avatar_url
    conversation_read.avatar_updated_at = avatar_updated_at
    return conversation_read


@app.patch("/api/conversations/{conversation_id}", response_model=ConversationRead)
def update_conversation(
    conversation_id: int,
    payload: ConversationUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ConversationRead:
    conversation = session.get(Conversation, conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa nao encontrada")
    ensure_conversation_access(conversation, current_user)
    update_data = payload.model_dump(exclude_unset=True)
    if "debtor_name" in update_data:
        new_name = (update_data["debtor_name"] or "").strip()
        conversation.debtor_name = new_name or (conversation.debtor_phone or "")
    conversation.updated_at = datetime.utcnow()
    session.add(conversation)
    session.commit()
    session.refresh(conversation)
    conversation_read = ConversationRead.model_validate(conversation)
    conversation_read.tags = visible_tags_for_user(conversation.tags, current_user)
    return conversation_read


@app.post("/api/conversations/{conversation_id}/pin", status_code=204)
def pin_conversation(
    conversation_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    conversation = session.get(Conversation, conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa nao encontrada")
    ensure_conversation_access(conversation, current_user)

    existing = session.get(ConversationPin, (conversation_id, current_user.id))
    if existing:
        return

    pin = ConversationPin(conversation_id=conversation_id, user_id=current_user.id)
    session.add(pin)
    session.commit()


@app.delete("/api/conversations/{conversation_id}/pin", status_code=204)
def unpin_conversation(
    conversation_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    conversation = session.get(Conversation, conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa nao encontrada")
    ensure_conversation_access(conversation, current_user)

    existing = session.get(ConversationPin, (conversation_id, current_user.id))
    if not existing:
        return

    session.delete(existing)
    session.commit()


@app.post("/api/conversations/{conversation_id}/mark-unread", status_code=204)
def mark_conversation_unread(
    conversation_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    conversation = session.get(Conversation, conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa nao encontrada")
    ensure_conversation_access(conversation, current_user)

    latest_debtor_message = session.exec(
        select(Message)
        .where(Message.conversation_id == conversation_id)
        .where(Message.direction == MessageDirection.debtor)
        .order_by(Message.timestamp.desc())
    ).first()
    if not latest_debtor_message:
        return
    if latest_debtor_message.is_read is False:
        return

    latest_debtor_message.is_read = False
    session.add(latest_debtor_message)
    session.commit()


@app.get("/api/conversations/{conversation_id}/messages", response_model=List[MessageRead])
def list_messages(
    conversation_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> List[MessageRead]:
    conversation = session.get(Conversation, conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa nao encontrada")
    ensure_conversation_access(conversation, current_user)
    messages = session.exec(
        select(Message)
        .where(Message.conversation_id == conversation_id)
        .order_by(Message.timestamp.asc())
    ).all()
    unread_messages = [
        message
        for message in messages
        if message.direction == MessageDirection.debtor and not message.is_read
    ]
    if unread_messages:
        for message in unread_messages:
            message.is_read = True
            session.add(message)
    session.commit()
    message_ids = [message.id for message in messages if message.id]
    deletion_map: Dict[int, MessageDeletion] = {}
    if message_ids:
        deletions = session.exec(
            select(MessageDeletion).where(MessageDeletion.message_id.in_(message_ids))
        ).all()
        deletion_map = {
            entry.message_id: entry for entry in deletions if entry.message_id is not None
        }
    reaction_counts_map: Dict[int, Dict[str, int]] = {}
    my_reaction_map: Dict[int, str] = {}
    if message_ids:
        reactions = session.exec(
            select(MessageReaction).where(MessageReaction.message_id.in_(message_ids))
        ).all()
        for reaction in reactions:
            if reaction.message_id is None:
                continue
            emoji = (reaction.emoji or "").strip()
            if not emoji:
                continue
            bucket = reaction_counts_map.setdefault(reaction.message_id, {})
            bucket[emoji] = bucket.get(emoji, 0) + 1
            if reaction.user_id == current_user.id:
                my_reaction_map[reaction.message_id] = emoji

    return [
        serialize_message_entity(
            message,
            deletion_map.get(message.id or 0),
            my_reaction=my_reaction_map.get(message.id or 0),
            reaction_counts=reaction_counts_map.get(message.id or 0, {}),
        )
        for message in messages
    ]


@app.post("/api/messages/{message_id}/delete-for-all", response_model=MessageRead)
def delete_message_for_all(
    message_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> MessageRead:
    message = session.get(Message, message_id)
    if not message:
        raise HTTPException(status_code=404, detail="Mensagem não encontrada.")
    if not message.conversation_id:
        raise HTTPException(status_code=400, detail="Mensagem sem conversa associada.")

    conversation = session.get(Conversation, message.conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa não encontrada.")
    ensure_conversation_access(conversation, current_user)

    if message.direction != MessageDirection.agent:
        raise HTTPException(status_code=400, detail="Só é possível apagar mensagens enviadas por você.")
    if not message.integration_message_id:
        raise HTTPException(status_code=400, detail="Mensagem sem ID do provedor para exclusão.")

    integration_session: Optional[DeviceSession] = None
    if message.via_session_id:
        integration_session = session.get(DeviceSession, message.via_session_id)
    if not integration_session:
        integration_session = resolve_uazapi_session_for_conversation(
            session, conversation=conversation, current_user=current_user
        )
    client = get_gateway_client(integration_session)
    if not isinstance(client, UazapiClient):
        raise HTTPException(status_code=400, detail="Apagar para todos disponível apenas para Uazapi.")
    if not client.is_configured:
        raise HTTPException(status_code=400, detail="Uazapi não configurada para este número.")

    try:
        client.delete_message_for_all(message.integration_message_id)
    except Exception as exc:
        logger.exception("Falha ao apagar mensagem na Uazapi: %s", exc)
        raise HTTPException(status_code=502, detail="Falha ao apagar mensagem na Uazapi.")

    deletion = session.get(MessageDeletion, message_id)
    if not deletion:
        deletion = MessageDeletion(message_id=message_id)
    deletion.deleted_for_all = True
    deletion.deleted_at = datetime.utcnow()
    deletion.deleted_by_user_id = current_user.id
    session.add(deletion)
    session.commit()
    session.refresh(deletion)

    return serialize_message_entity(message, deletion)


@app.post("/api/messages/{message_id}/edit", response_model=MessageRead)
def edit_message(
    message_id: int,
    payload: MessageEditRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> MessageRead:
    message = session.get(Message, message_id)
    if not message:
        raise HTTPException(status_code=404, detail="Mensagem não encontrada.")
    if not message.conversation_id:
        raise HTTPException(status_code=400, detail="Mensagem sem conversa associada.")

    conversation = session.get(Conversation, message.conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa não encontrada.")
    ensure_conversation_access(conversation, current_user)

    if message.direction != MessageDirection.agent:
        raise HTTPException(status_code=400, detail="Só é possível editar mensagens enviadas por você.")
    if message.message_type != MessageType.text:
        raise HTTPException(status_code=400, detail="Só é possível editar mensagens de texto.")
    if not payload.text or not payload.text.strip():
        raise HTTPException(status_code=400, detail="Texto inválido.")

    if not current_user.is_admin and message.author_user_id and message.author_user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Você não pode editar esta mensagem.")

    if not message.integration_message_id:
        raise HTTPException(status_code=400, detail="Mensagem sem ID do provedor para edição.")

    integration_session: Optional[DeviceSession] = None
    if message.via_session_id:
        integration_session = session.get(DeviceSession, message.via_session_id)
    if not integration_session:
        integration_session = resolve_uazapi_session_for_conversation(
            session, conversation=conversation, current_user=current_user
        )
    client = get_gateway_client(integration_session)
    if not isinstance(client, UazapiClient):
        raise HTTPException(status_code=400, detail="Editar mensagem disponível apenas para Uazapi.")
    if not client.is_configured:
        raise HTTPException(status_code=400, detail="Uazapi não configurada para este número.")

    new_text = payload.text.strip()
    try:
        provider_response = client.edit_message(message_id=message.integration_message_id, text=new_text)
    except Exception as exc:
        logger.exception("Falha ao editar mensagem na Uazapi: %s", exc)
        raise HTTPException(status_code=502, detail="Falha ao editar mensagem na Uazapi.")

    provider_message_id = extract_provider_message_id(provider_response)
    if not message.original_content:
        message.original_content = message.content
    message.content = new_text
    message.edited_at = datetime.utcnow()
    message.edited_by_user_id = current_user.id
    if provider_message_id:
        message.integration_message_id = provider_message_id

    session.add(message)
    session.commit()
    session.refresh(message)

    deletion_entry = session.get(MessageDeletion, message_id)
    reactions = session.exec(select(MessageReaction).where(MessageReaction.message_id == message_id)).all()
    reaction_counts: Dict[str, int] = {}
    my_reaction: Optional[str] = None
    for reaction in reactions:
        emoji = (reaction.emoji or "").strip()
        if not emoji:
            continue
        reaction_counts[emoji] = reaction_counts.get(emoji, 0) + 1
        if reaction.user_id == current_user.id:
            my_reaction = emoji

    return serialize_message_entity(
        message,
        deletion_entry,
        my_reaction=my_reaction,
        reaction_counts=reaction_counts,
    )


@app.post("/api/messages/{message_id}/react", response_model=MessageRead)
def react_to_message(
    message_id: int,
    payload: MessageReactRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> MessageRead:
    message = session.get(Message, message_id)
    if not message:
        raise HTTPException(status_code=404, detail="Mensagem não encontrada.")
    if not message.conversation_id:
        raise HTTPException(status_code=400, detail="Mensagem sem conversa associada.")
    conversation = session.get(Conversation, message.conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa não encontrada.")
    ensure_conversation_access(conversation, current_user)

    if not message.integration_message_id:
        raise HTTPException(
            status_code=400, detail="Mensagem sem ID do provedor para reagir."
        )
    if not conversation.debtor_phone:
        raise HTTPException(status_code=400, detail="Conversa sem telefone associado.")

    integration_session: Optional[DeviceSession] = None
    if message.via_session_id:
        integration_session = session.get(DeviceSession, message.via_session_id)
    if not integration_session:
        integration_session = resolve_uazapi_session_for_conversation(
            session, conversation=conversation, current_user=current_user
        )
    client = get_gateway_client(integration_session)
    if not isinstance(client, UazapiClient):
        raise HTTPException(status_code=400, detail="Reações disponíveis apenas para Uazapi.")
    if not client.is_configured:
        raise HTTPException(status_code=400, detail="Uazapi não configurada para este número.")

    emoji = (payload.emoji or "").strip()

    try:
        client.react_message(conversation.debtor_phone, message.integration_message_id, emoji)
    except Exception as exc:
        logger.exception("Falha ao reagir mensagem na Uazapi: %s", exc)
        raise HTTPException(status_code=502, detail="Falha ao reagir mensagem na Uazapi.")

    existing = session.get(MessageReaction, (message_id, current_user.id))
    if not emoji:
        if existing:
            session.delete(existing)
            session.commit()
    else:
        if not existing:
            existing = MessageReaction(message_id=message_id, user_id=current_user.id)
        existing.emoji = emoji
        existing.reacted_at = datetime.utcnow()
        session.add(existing)
        session.commit()

    deletion = session.get(MessageDeletion, message_id)
    reactions = session.exec(
        select(MessageReaction).where(MessageReaction.message_id == message_id)
    ).all()
    counts: Dict[str, int] = {}
    my_reaction: Optional[str] = None
    for reaction in reactions:
        em = (reaction.emoji or "").strip()
        if not em:
            continue
        counts[em] = counts.get(em, 0) + 1
        if reaction.user_id == current_user.id:
            my_reaction = em
    return serialize_message_entity(
        message,
        deletion,
        my_reaction=my_reaction,
        reaction_counts=counts,
    )


@app.get(
    "/api/conversations/{conversation_id}/notes",
    response_model=List[ConversationNoteRead],
)
def list_conversation_notes(
    conversation_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> List[ConversationNoteRead]:
    conversation = session.get(Conversation, conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa nao encontrada")
    ensure_conversation_access(conversation, current_user)
    notes = session.exec(
        select(ConversationNote)
        .where(ConversationNote.conversation_id == conversation_id)
        .order_by(ConversationNote.created_at.asc())
    ).all()
    author_ids = {note.created_by_user_id for note in notes if note.created_by_user_id}
    author_map: Dict[int, User] = {}
    if author_ids:
        authors = session.exec(select(User).where(User.id.in_(author_ids))).all()
        author_map = {author.id: author for author in authors if author}
    response: List[ConversationNoteRead] = []
    for note in notes:
        note_data = ConversationNoteRead.model_validate(note)
        author = author_map.get(note.created_by_user_id)
        if author:
            note_data.author_name = author.full_name or author.username
        response.append(note_data)
    return response


@app.post(
    "/api/conversations/{conversation_id}/notes",
    response_model=ConversationNoteRead,
)
def create_conversation_note(
    conversation_id: int,
    payload: ConversationNoteCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ConversationNoteRead:
    conversation = session.get(Conversation, conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa nao encontrada")
    ensure_conversation_access(conversation, current_user)
    if not current_user.is_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Apenas administradores podem criar notas internas",
        )
    text = (payload.text or "").strip()
    if not text:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Informe o texto da nota",
        )
    note = ConversationNote(
        organization_id=current_user.organization_id,
        conversation_id=conversation.id,
        created_by_user_id=current_user.id,
        text=text,
    )
    session.add(note)
    session.commit()
    session.refresh(note)
    note_data = ConversationNoteRead.model_validate(note)
    note_data.author_name = current_user.full_name or current_user.username
    return note_data


@app.get("/api/conversations/{conversation_id}/avatar")
def get_conversation_avatar(
    conversation_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> Dict[str, Optional[str]]:
    conversation = session.get(Conversation, conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa nao encontrada")
    ensure_conversation_access(conversation, current_user)
    avatar_entry = session.get(ConversationAvatar, conversation_id)
    now = datetime.utcnow()
    if avatar_entry and avatar_entry.media_path and avatar_entry.updated_at:
        if avatar_entry.updated_at > now - AVATAR_CACHE_TTL:
            url = build_public_media_url(avatar_entry.media_path, request)
            if url:
                return {"url": url}
    if avatar_entry and not avatar_entry.media_path:
        if avatar_entry.updated_at and avatar_entry.updated_at > now - AVATAR_CACHE_TTL:
            return {"url": None}
    target_user_id = conversation.owner_user_id or current_user.id
    if target_user_id != current_user.id and not current_user.is_admin:
        target_user_id = current_user.id
    integration_session = ensure_session_exists(
        session, current_user.organization_id, target_user_id
    )
    url = fetch_and_store_conversation_avatar(
        session,
        conversation,
        integration_session,
        request,
        preview=True,
    )
    return {"url": url}


@app.post("/api/messages", response_model=MessageRead)
def send_message(
    payload: MessageCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> MessageRead:
    conversation = session.get(Conversation, payload.conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa invalida")
    ensure_conversation_access(conversation, current_user)

    quoted_message = ensure_reply_target(session, conversation.id, payload.reply_to_message_id)
    integration_session = ensure_session_exists(
        session, current_user.organization_id, current_user.id
    )
    message = Message(
        conversation_id=payload.conversation_id,
        direction=payload.direction,
        content=payload.content,
        reply_to_message_id=payload.reply_to_message_id,
        via_session_id=integration_session.id,
        is_read=payload.direction == MessageDirection.agent,
        author_user_id=current_user.id,
        message_type=payload.message_type,
        media_path=payload.media_path,
        media_content_type=payload.media_content_type,
        media_size_bytes=payload.media_size_bytes,
        media_duration_seconds=payload.media_duration_seconds,
    )
    session.add(message)
    conversation.updated_at = datetime.utcnow()
    session.add(conversation)
    session.commit()
    session.refresh(message)

    provider_response: Optional[Any] = None
    if payload.direction == MessageDirection.agent:
        effective_client = get_gateway_client(integration_session)
        if not effective_client.is_configured:
            provider_label = (integration_session.provider if integration_session else "Evolution") or "Evolution"
            if provider_label.lower() == "evolution":
                detail = "Integracao Evolution nao configurada. Informe a instância e token em Integração > Gerar novo QR."
            else:
                detail = f"Integracao {provider_label} nao configurada. Informe URL/instancia/token em Integração > Gerar novo QR."
            raise HTTPException(status_code=400, detail=detail)
        quoted_gateway_id = quoted_message.integration_message_id if quoted_message else None
        try:
            provider_response = effective_client.send_text_message(
                conversation.debtor_phone,
                payload.content,
                quoted_message_id=quoted_gateway_id,
            )
            if integration_session.status != DeviceSessionStatus.connected:
                integration_session.status = DeviceSessionStatus.connected
            integration_session.last_synced_at = datetime.utcnow()
            session.add(integration_session)
            session.commit()
            session.refresh(integration_session)
        except Exception as exc:  # pragma: no cover
            provider_label = (integration_session.provider if integration_session else "Evolution") or "Evolution"
            raise HTTPException(
                status_code=502,
                detail=f"Erro ao enviar mensagem para {provider_label}: {exc}",
            ) from exc

    provider_message_id = extract_provider_message_id(provider_response)
    if provider_message_id:
        message.integration_message_id = provider_message_id
        session.add(message)
        session.commit()
        session.refresh(message)

    sse_notify(
        current_user.organization_id,
        "message",
        {
            "conversation_id": conversation.id,
            "message_id": message.id,
            "direction": getattr(message.direction, "value", str(message.direction)),
        },
    )

    return serialize_message_entity(message)


@app.post("/api/messages/audio", response_model=MessageRead)
async def send_audio_message(
    request: Request,
    conversation_id: int = Form(...),
    duration: Optional[float] = Form(None),
    caption: Optional[str] = Form(None),
    reply_to_message_id: Optional[int] = Form(None),
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> MessageRead:
    conversation = session.get(Conversation, conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa invalida")
    ensure_conversation_access(conversation, current_user)
    integration_session = ensure_session_exists(
        session, current_user.organization_id, current_user.id
    )
    quoted_message = quoted_message = ensure_reply_target(session, conversation.id, reply_to_message_id)
    raw_data = await file.read()
    if not raw_data:
        raise HTTPException(status_code=400, detail="Arquivo de áudio vazio")
    processed_data, processed_filename, processed_content_type, was_converted = ensure_voice_note_audio(
        raw_data,
        filename=file.filename or "audio-message.ogg",
        content_type=file.content_type or "audio/ogg",
    )
    if was_converted:
        logger.info("Audio convertido para OGG/Opus: %s -> %s", file.filename, processed_filename)
    stored_media = media_storage.save_bytes(
        processed_data,
        filename=processed_filename,
        content_type=processed_content_type,
        prefix=build_media_prefix(conversation),
    )
    public_media_url: Optional[str] = None
    if stored_media.key:
        candidate_url: Optional[str] = None
        try:
            candidate_url = media_storage.build_url(stored_media.key)
        except Exception:
            candidate_url = None
        if candidate_url:
            if candidate_url.startswith("http://") or candidate_url.startswith("https://"):
                public_media_url = candidate_url
            else:
                base = str(request.base_url).rstrip("/") if request else ""
                public_media_url = f"{base}{candidate_url}"
        if not public_media_url and media_storage.is_local():
            try:
                public_media_url = str(request.url_for("media", path=stored_media.key))
            except Exception:
                public_media_url = None
    message = Message(
        conversation_id=conversation.id,
        direction=MessageDirection.agent,
        content=caption or "",
        reply_to_message_id=reply_to_message_id,
        via_session_id=integration_session.id,
        is_read=True,
        author_user_id=current_user.id,
        message_type=MessageType.audio,
        media_path=stored_media.key,
        media_content_type=stored_media.content_type,
        media_size_bytes=stored_media.size,
        media_duration_seconds=duration,
    )
    session.add(message)
    conversation.updated_at = datetime.utcnow()
    session.add(conversation)
    session.commit()
    session.refresh(message)

    effective_client = get_gateway_client(integration_session)
    if not effective_client.is_configured:
        raise HTTPException(
            status_code=400,
            detail="Integracao Evolution nao configurada. Informe a instƒncia e token em Integra‡Æo > Gerar novo QR.",
        )
    send_as_voice_note = False
    if processed_content_type:
        lowered_type = processed_content_type.split(";", 1)[0].strip().lower()
        if lowered_type in {"audio/ogg", "application/ogg", "audio/opus"}:
            send_as_voice_note = True
    provider_response: Optional[Any] = None
    try:
        provider_response = effective_client.send_audio_message(
            conversation.debtor_phone,
            processed_data,
            filename=processed_filename,
            mime_type=stored_media.content_type or processed_content_type,
            media_url=public_media_url,
            as_voice_note=send_as_voice_note,
            quoted_message_id=quoted_message.integration_message_id if quoted_message else None,
        )
        if integration_session.status != DeviceSessionStatus.connected:
            integration_session.status = DeviceSessionStatus.connected
        integration_session.last_synced_at = datetime.utcnow()
        session.add(integration_session)
        session.commit()
        session.refresh(integration_session)
    except requests.HTTPError as exc:  # pragma: no cover
        provider_label = (integration_session.provider if integration_session else "Evolution") or "Evolution"
        body = ""
        if getattr(exc, "response", None) is not None:
            try:
                body = exc.response.text[:500]  # type: ignore[attr-defined]
            except Exception:  # defensive
                body = ""
        logger.error("Envio de audio para %s falhou: %s - body=%s", provider_label, exc, body)
        raise HTTPException(
            status_code=502,
            detail=f"Erro ao enviar áudio para {provider_label}: {exc} - body: {body}",
        ) from exc
    except Exception as exc:  # pragma: no cover
        provider_label = (integration_session.provider if integration_session else "Evolution") or "Evolution"
        raise HTTPException(
            status_code=502,
            detail=f"Erro ao enviar áudio para {provider_label}: {exc}",
    ) from exc

    provider_message_id = extract_provider_message_id(provider_response)
    if provider_message_id:
        message.integration_message_id = provider_message_id
        session.add(message)
        session.commit()
        session.refresh(message)

    sse_notify(
        current_user.organization_id,
        "message",
        {
            "conversation_id": conversation.id,
            "message_id": message.id,
            "direction": getattr(message.direction, "value", str(message.direction)),
        },
    )

    return serialize_message_entity(message)


@app.post("/api/messages/media", response_model=MessageRead)
async def send_media_file(
    request: Request,
    conversation_id: int = Form(...),
    caption: Optional[str] = Form(None),
    reply_to_message_id: Optional[int] = Form(None),
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> MessageRead:
    conversation = session.get(Conversation, conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa invalida")
    ensure_conversation_access(conversation, current_user)
    integration_session = ensure_session_exists(
        session, current_user.organization_id, current_user.id
    )
    quoted_message = ensure_reply_target(session, conversation.id, reply_to_message_id)
    raw_data = await file.read()
    if not raw_data:
        raise HTTPException(status_code=400, detail="Arquivo vazio")
    message_type, sanitized_type = infer_uploaded_media_type(file.filename, file.content_type)
    stored_media = media_storage.save_bytes(
        raw_data,
        filename=file.filename or "arquivo.bin",
        content_type=sanitized_type,
        prefix=build_media_prefix(conversation),
    )
    public_media_url: Optional[str] = None
    if stored_media.key:
        candidate_url: Optional[str] = None
        try:
            candidate_url = media_storage.build_url(stored_media.key)
        except Exception:
            candidate_url = None
        if candidate_url:
            if candidate_url.startswith("http://") or candidate_url.startswith("https://"):
                public_media_url = candidate_url
            else:
                base = str(request.base_url).rstrip("/") if request else ""
                public_media_url = f"{base}{candidate_url}"
        if not public_media_url and media_storage.is_local():
            try:
                public_media_url = str(request.url_for("media", path=stored_media.key))
            except Exception:
                public_media_url = None
    message = Message(
        conversation_id=conversation.id,
        direction=MessageDirection.agent,
        content=caption or "",
        reply_to_message_id=reply_to_message_id,
        via_session_id=integration_session.id,
        is_read=True,
        author_user_id=current_user.id,
        message_type=message_type,
        media_path=stored_media.key,
        media_content_type=stored_media.content_type,
        media_size_bytes=stored_media.size,
    )
    session.add(message)
    conversation.updated_at = datetime.utcnow()
    session.add(conversation)
    session.commit()
    session.refresh(message)

    effective_client = get_gateway_client(integration_session)
    if not effective_client.is_configured:
        raise HTTPException(
            status_code=400,
            detail="Integracao nao configurada. Informe URL/instancia/token em Integra‡Æo > Gerar novo QR.",
        )
    try:
        send_media = getattr(effective_client, "send_media_file", None)
        if not callable(send_media):
            raise HTTPException(status_code=400, detail="Gateway nao suporta envio de arquivos.")
        send_media(
            conversation.debtor_phone,
            raw_data,
            filename=file.filename or stored_media.key or "arquivo.bin",
            mime_type=sanitized_type,
            media_category=message_type.value,
            caption=caption or "",
            media_url=public_media_url,
            quoted_message_id=quoted_message.integration_message_id if quoted_message else None,
        )
        if integration_session.status != DeviceSessionStatus.connected:
            integration_session.status = DeviceSessionStatus.connected
        integration_session.last_synced_at = datetime.utcnow()
        session.add(integration_session)
        session.commit()
        session.refresh(integration_session)
    except HTTPException:
        raise
    except requests.HTTPError as exc:  # pragma: no cover
        body = ""
        if getattr(exc, "response", None) is not None:
            try:
                body = exc.response.text[:500]  # type: ignore[attr-defined]
            except Exception:
                body = ""
        logger.error("Envio de arquivo falhou: %s - body=%s", exc, body)
        raise HTTPException(
            status_code=502,
            detail=f"Erro ao enviar arquivo para gateway: {exc} - body: {body}",
        ) from exc
    except Exception as exc:  # pragma: no cover
        raise HTTPException(
            status_code=502,
            detail=f"Erro ao enviar arquivo: {exc}",
        ) from exc

    sse_notify(
        current_user.organization_id,
        "message",
        {
            "conversation_id": conversation.id,
            "message_id": message.id,
            "direction": getattr(message.direction, "value", str(message.direction)),
        },
    )

    return serialize_message_entity(message)


@app.post("/api/calls/start", response_model=CallActionResponse)
def start_voice_call(
    payload: CallStartRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> CallActionResponse:
    conversation = session.get(Conversation, payload.conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa nao encontrada")
    ensure_conversation_access(conversation, current_user)
    phone_number = payload.phone_number or conversation.debtor_phone
    if not phone_number:
        raise HTTPException(status_code=400, detail="Nao ha telefone associado a esta conversa.")
    integration_session = ensure_session_exists(
        session, current_user.organization_id, current_user.id
    )
    ensure_voice_call_capability(integration_session)
    if not session_has_credentials(integration_session):
        raise HTTPException(
            status_code=400,
            detail="Integracao Uazapi nao configurada. Informe URL/instancia/token em Integracao > Gerar novo QR.",
        )
    client = get_gateway_client(integration_session)
    starter = getattr(client, "start_voice_call", None)
    if not callable(starter):
        raise HTTPException(
            status_code=400,
            detail="Gateway selecionado nao suporta chamadas de voz.",
        )
    try:
        provider_response = starter(phone_number)
    except requests.HTTPError as exc:  # pragma: no cover - integracao externa
        body = ""
        if getattr(exc, "response", None) is not None:
            try:
                body = exc.response.text[:500]  # type: ignore[attr-defined]
            except Exception:
                body = ""
        raise HTTPException(
            status_code=502,
            detail=f"Erro ao iniciar chamada na Uazapi: {body or exc}",
        ) from exc
    except Exception as exc:  # pragma: no cover - integracao externa
        raise HTTPException(
            status_code=502,
            detail=f"Erro inesperado ao iniciar chamada: {exc}",
        ) from exc
    integration_session.status = DeviceSessionStatus.connected
    integration_session.last_synced_at = datetime.utcnow()
    session.add(integration_session)
    session.commit()
    session.refresh(integration_session)
    detail = ""
    if isinstance(provider_response, dict):
        detail = provider_response.get("response") or provider_response.get("message") or ""
    if not detail:
        detail = "Chamada iniciada com sucesso."
    return CallActionResponse(
        success=True,
        detail=detail,
        provider_response=provider_response if isinstance(provider_response, dict) else None,
    )


@app.post("/api/calls/reject", response_model=CallActionResponse)
def reject_voice_call(
    payload: CallRejectRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> CallActionResponse:
    phone_number = payload.phone_number
    if payload.conversation_id:
        conversation = session.get(Conversation, payload.conversation_id)
        if not conversation:
            raise HTTPException(status_code=404, detail="Conversa nao encontrada")
        ensure_conversation_access(conversation, current_user)
        phone_number = phone_number or conversation.debtor_phone
    if not phone_number and not payload.call_id:
        raise HTTPException(
            status_code=400,
            detail="Informe um numero ou ID da chamada para rejeitar.",
        )
    integration_session = ensure_session_exists(
        session, current_user.organization_id, current_user.id
    )
    ensure_voice_call_capability(integration_session)
    if not session_has_credentials(integration_session):
        raise HTTPException(
            status_code=400,
            detail="Integracao Uazapi nao configurada. Informe URL/instancia/token em Integracao > Gerar novo QR.",
        )
    client = get_gateway_client(integration_session)
    rejector = getattr(client, "reject_voice_call", None)
    if not callable(rejector):
        raise HTTPException(
            status_code=400,
            detail="Gateway selecionado nao suporta rejeicao de chamadas.",
        )
    try:
        provider_response = rejector(phone_number=phone_number, call_id=payload.call_id)
    except requests.HTTPError as exc:  # pragma: no cover - integracao externa
        body = ""
        if getattr(exc, "response", None) is not None:
            try:
                body = exc.response.text[:500]  # type: ignore[attr-defined]
            except Exception:
                body = ""
        raise HTTPException(
            status_code=502,
            detail=f"Erro ao rejeitar chamada na Uazapi: {body or exc}",
        ) from exc
    except Exception as exc:  # pragma: no cover - integracao externa
        raise HTTPException(
            status_code=502,
            detail=f"Erro inesperado ao rejeitar chamada: {exc}",
        ) from exc
    integration_session.last_synced_at = datetime.utcnow()
    session.add(integration_session)
    session.commit()
    session.refresh(integration_session)
    detail = ""
    if isinstance(provider_response, dict):
        detail = provider_response.get("response") or provider_response.get("message") or ""
    if not detail:
        detail = "Chamada rejeitada com sucesso."
    return CallActionResponse(
        success=True,
        detail=detail,
        provider_response=provider_response if isinstance(provider_response, dict) else None,
    )


@app.post("/api/messages/incoming", response_model=MessageRead)
def simulate_incoming_message(
    payload: IncomingMessagePayload,
    session: Session = Depends(get_session),
) -> MessageRead:
    organization_id: Optional[int] = None
    conversation = None
    session_owner_id: Optional[int] = None
    normalized_incoming_phone = normalize_phone(payload.debtor_phone)
    normalized_display_name = normalize_display_name(payload.debtor_name)
    if payload.conversation_id:
        conversation = session.get(Conversation, payload.conversation_id)
        if conversation:
            organization_id = conversation.organization_id
            if not conversation.owner_user_id:
                owner = get_default_owner(session, organization_id)
                conversation.owner_user_id = owner.id
                session.add(conversation)
                session.commit()
                session.refresh(conversation)
    elif payload.session_code:
        integration_session = session.exec(
            select(DeviceSession).where(DeviceSession.session_code == payload.session_code)
        ).first()
        if integration_session:
            organization_id = integration_session.organization_id
            session_owner_id = integration_session.owner_user_id
            integration_session.status = DeviceSessionStatus.connected
            integration_session.last_synced_at = datetime.utcnow()
            session.add(integration_session)
            session.commit()
            session.refresh(integration_session)
    if not conversation and payload.debtor_phone and organization_id:
        candidates = phone_variants(payload.debtor_phone)
        if candidates:
            query = (
                select(Conversation)
                .where(Conversation.organization_id == organization_id)
                .where(Conversation.debtor_phone.in_(candidates))
            )
            if session_owner_id:
                query = query.where(Conversation.owner_user_id == session_owner_id)
            conversation = session.exec(query).first()
        if conversation and not conversation.owner_user_id:
            owner = get_default_owner(session, organization_id)
            conversation.owner_user_id = owner.id
            session.add(conversation)
            session.commit()
            session.refresh(conversation)
    if not conversation:
        if not organization_id:
            raise HTTPException(status_code=400, detail="Sessao ou conversa invalida")
        owner = (
            session.get(User, session_owner_id) if session_owner_id else get_default_owner(session, organization_id)
        )
        conversation = Conversation(
            debtor_name=normalized_display_name or payload.debtor_phone or "Contato",
            debtor_phone=normalized_incoming_phone or (payload.debtor_phone or ""),
            organization_id=organization_id,
            owner_user_id=owner.id,
        )
        session.add(conversation)
        session.commit()
        session.refresh(conversation)
    else:
        if normalized_display_name:
            current_name = (conversation.debtor_name or "").strip()
            current_digits = clean_phone_digits(current_name) if current_name else None
            incoming_digits = normalized_incoming_phone or clean_phone_digits(payload.debtor_phone)
            if (
                not current_name
                or current_name == (payload.debtor_phone or "")
                or (current_digits and incoming_digits and current_digits == incoming_digits)
            ):
                conversation.debtor_name = normalized_display_name

    owner_id = conversation.owner_user_id
    if not owner_id:
        owner = get_default_owner(session, conversation.organization_id)
        conversation.owner_user_id = owner.id
        session.add(conversation)
        session.commit()
        session.refresh(conversation)
        owner_id = owner.id

    if payload.media_blob:
        fallback_type = payload.media_content_type or "application/octet-stream"
        stored_name = determine_incoming_filename(
            payload.media_original_name,
            payload.message_type,
            fallback_type,
        )
        stored_media = media_storage.save_bytes(
            payload.media_blob,
            filename=stored_name,
            content_type=fallback_type,
            prefix=build_media_prefix(conversation),
        )
        payload.media_path = stored_media.key
        payload.media_content_type = stored_media.content_type
        payload.media_size_bytes = stored_media.size
        payload.media_blob = None

    reply_target_id: Optional[int] = None
    if payload.quoted_integration_message_id:
        quoted_message = session.exec(
            select(Message)
            .where(Message.conversation_id == conversation.id)
            .where(Message.integration_message_id == payload.quoted_integration_message_id)
            .order_by(Message.id.desc())
        ).first()
        if quoted_message:
            reply_target_id = quoted_message.id

    message = Message(
        conversation_id=conversation.id,
        direction=MessageDirection.debtor,
        content=payload.content or "",
        is_read=False,
        author_user_id=owner_id,
        message_type=payload.message_type,
        media_path=payload.media_path,
        media_content_type=payload.media_content_type,
        media_size_bytes=payload.media_size_bytes,
        media_duration_seconds=payload.media_duration_seconds,
        integration_message_id=payload.integration_message_id,
        reply_to_message_id=reply_target_id,
    )
    conversation.updated_at = datetime.utcnow()
    session.add(message)
    session.add(conversation)
    session.commit()
    session.refresh(message)
    sse_notify(
        conversation.organization_id,
        "message",
        {
            "conversation_id": conversation.id,
            "message_id": message.id,
            "direction": getattr(message.direction, "value", message.direction),
        },
    )
    return serialize_message_entity(message)


@app.post("/webhooks/evolution")
async def evolution_webhook(
    request: Request,
    session: Session = Depends(get_session),
) -> Dict[str, Any]:
    try:
        raw_body = await request.json()
    except Exception as exc:  # pragma: no cover - defensive
        raise HTTPException(status_code=400, detail="JSON inválido") from exc
    append_webhook_dump("evolution", raw_body)

    events: List[Dict[str, Any]]
    if isinstance(raw_body, list):
        events = [item for item in raw_body if isinstance(item, dict)]
    elif isinstance(raw_body, dict):
        events = [raw_body]
    else:
        events = []

    processed = 0
    for event in events:
        event_name = event.get("event") or event.get("type")
        if event_name not in {"messages.upsert", "message"}:
            continue
        instance_identifier = (
            event.get("instance")
            or event.get("instance_id")
            or event.get("instanceId")
            or event.get("session")
        )
        integration_session = find_session_by_instance(session, instance_identifier)
        if not integration_session:
            logger.warning("Webhook ignorado; instancia desconhecida: %s", instance_identifier)
            continue
        data = event.get("data") or []
        if isinstance(data, dict):
            data = [data]
        for raw_message in data:
            if not isinstance(raw_message, dict):
                continue
            key = raw_message.get("key") or {}
            if key.get("fromMe"):
                continue
            message_payload = raw_message.get("message") or {}
            text = extract_evolution_text(message_payload)
            audio_payload = extract_evolution_audio(message_payload)
            document_payload = extract_evolution_document(message_payload)
            if not text and not audio_payload and not document_payload:
                continue
            debtor_phone = normalize_whatsapp_jid(key.get("remoteJid"))
            if not debtor_phone:
                continue
            blob_bytes: Optional[bytes] = None
            media_name: Optional[str] = None
            media_type: Optional[str] = None
            media_size: Optional[int] = None
            media_duration: Optional[float] = None
            message_type = (
                MessageType.audio if audio_payload else
                (document_payload[1] if document_payload else MessageType.text)
            )
            if audio_payload:
                media_duration = (
                    audio_payload.get("seconds")
                    or audio_payload.get("duration")
                    or audio_payload.get("length")
                )
                try:
                    downloaded = download_media_from_evolution(audio_payload, integration_session)
                except Exception as exc:  # pragma: no cover - defensive
                    logger.warning("Erro ao baixar audio da Evolution: %s", exc)
                    downloaded = None
                if downloaded:
                    blob_bytes, media_type, media_name = downloaded
                    media_size = len(blob_bytes)
                if not text:
                    text = audio_payload.get("caption") or "[Áudio]"
            elif document_payload:
                payload_data, payload_type = document_payload
                try:
                    downloaded = download_media_from_evolution(payload_data, integration_session)
                except Exception as exc:  # pragma: no cover - defensive
                    logger.warning("Erro ao baixar documento da Evolution: %s", exc)
                    downloaded = None
                if downloaded:
                    blob_bytes, media_type, media_name = downloaded
                    media_size = len(blob_bytes)
                else:
                    message_type = MessageType.text
            if not text:
                text = (
                    payload_data.get("caption")
                    or payload_data.get("title")
                    or payload_data.get("fileName")
                    or "[Arquivo recebido]"
                )
            quoted_reference = extract_evolution_reply_reference(raw_message)
            incoming_payload = IncomingMessagePayload(
                debtor_phone=debtor_phone,
                content=text or "",
                session_code=integration_session.session_code,
                integration_message_id=key.get("id"),
                message_type=message_type,
                media_blob=blob_bytes,
                media_original_name=media_name,
                media_content_type=media_type,
                media_size_bytes=media_size,
                media_duration_seconds=media_duration,
                quoted_integration_message_id=quoted_reference,
            )
            simulate_incoming_message(incoming_payload, session)
            processed += 1
    return {"processed": processed}


@app.post("/webhooks/uazapi")
async def uazapi_webhook(
    request: Request,
    session: Session = Depends(get_session),
) -> Dict[str, Any]:
    def extract_instance_identifier(*sources: Optional[Dict[str, Any]]) -> Optional[str]:
        for source in sources:
            if not isinstance(source, dict):
                continue
            for key in (
                "instance",
                "instance_id",
                "instanceId",
                "instanceName",
                "session",
                "session_id",
                "sessionId",
                "sessionName",
                "session_name",
                "sessionCode",
                "session_code",
            ):
                raw = source.get(key)
                if isinstance(raw, str) and raw.strip():
                    return raw.strip()
        return None

    MessageEntry = Tuple[Dict[str, Any], Optional[str]]

    def collect_message_entries(
        payload: Any,
        *,
        inherited_instance: Optional[str] = None,
    ) -> List[MessageEntry]:
        entries: List[MessageEntry] = []

        if isinstance(payload, list):
            for item in payload:
                entries.extend(
                    collect_message_entries(item, inherited_instance=inherited_instance)
                )
            return entries

        if not isinstance(payload, dict):
            return entries

        current_instance = extract_instance_identifier(payload) or inherited_instance

        inner_message = payload.get("message")
        if isinstance(inner_message, (dict, list)):
            entries.extend(
                collect_message_entries(inner_message, inherited_instance=current_instance)
            )
            return entries

        inner_messages = payload.get("messages")
        if isinstance(inner_messages, (dict, list)):
            entries.extend(
                collect_message_entries(inner_messages, inherited_instance=current_instance)
            )
            return entries

        inner_data = payload.get("data")
        if isinstance(inner_data, (dict, list)):
            entries.extend(
                collect_message_entries(inner_data, inherited_instance=current_instance)
            )
            return entries

        entries.append((payload, current_instance))
        return entries

    def resolve_uazapi_debtor_name(
        message_payload: Optional[Dict[str, Any]],
        chat_context: Optional[Dict[str, Any]],
    ) -> Optional[str]:
        candidates: List[Optional[str]] = []
        if isinstance(message_payload, dict):
            candidates.extend(
                [
                    message_payload.get("groupName"),
                    message_payload.get("chatName"),
                    message_payload.get("chatname"),
                ]
            )
            if not message_payload.get("isGroup"):
                candidates.extend(
                    [
                        message_payload.get("senderName"),
                        message_payload.get("pushName"),
                        message_payload.get("contactName"),
                    ]
                )
        if isinstance(chat_context, dict):
            candidates.extend(
                [
                    chat_context.get("name"),
                    chat_context.get("wa_name"),
                    chat_context.get("wa_contactName"),
                    chat_context.get("lead_fullName"),
                    chat_context.get("lead_name"),
                ]
            )
        for candidate in candidates:
            normalized = normalize_display_name(candidate)
            if normalized:
                return normalized
        return None

    try:
        raw_body = await request.json()
    except Exception as exc:  # pragma: no cover - defensive
        raise HTTPException(status_code=400, detail="JSON invalido") from exc

    append_webhook_dump("uazapi", raw_body)

    events: List[Dict[str, Any]]
    if isinstance(raw_body, list):
        events = [item for item in raw_body if isinstance(item, dict)]
    elif isinstance(raw_body, dict):
        events = [raw_body]
    else:
        events = []

    processed = 0
    for event in events:
        event_name = (
            event.get("event")
            or event.get("type")
            or event.get("EventType")
            or event.get("eventType")
            or ""
        ).lower()
        data_block = event.get("data")
        if data_block is None:
            if "message" in event:
                data_block = event.get("message")
            elif "messages" in event:
                data_block = event.get("messages")
            else:
                data_block = event
        event_instance_identifier = extract_instance_identifier(event, data_block)
        if event_name in {"connection", "status"}:
            instance_identifier = event_instance_identifier
            if not instance_identifier:
                continue
            integration_session = find_session_by_instance(session, instance_identifier)
            if not integration_session:
                logger.warning("Webhook Uazapi ignorado; instancia desconhecida: %s", instance_identifier)
                continue
            payload = data_block if isinstance(data_block, dict) else {}
            status_value = (payload.get("status") or payload.get("state") or "").lower()
            if status_value == "connected":
                integration_session.status = DeviceSessionStatus.connected
            elif status_value in {"disconnected", "closed", "erro", "error"}:
                integration_session.status = DeviceSessionStatus.disconnected
            integration_session.last_synced_at = datetime.utcnow()
            session.add(integration_session)
            session.commit()
            session.refresh(integration_session)
            continue
        if event_name not in {"message", "messages", "messages.upsert"}:
            continue

        message_entries = collect_message_entries(
            data_block,
            inherited_instance=event_instance_identifier,
        )
        if not message_entries:
            continue
        chat_context = event.get("chat") if isinstance(event.get("chat"), dict) else None

        for message_data, fallback_instance in message_entries:
            if not isinstance(message_data, dict):
                continue
            if message_data.get("fromMe"):
                continue
            is_group_message = bool(message_data.get("isGroup"))
            instance_identifier = extract_instance_identifier(message_data) or fallback_instance
            if not instance_identifier:
                logger.warning("Webhook Uazapi sem instancia; ignorado payload %s", message_data.keys())
                continue
            integration_session = find_session_by_instance(session, instance_identifier)
            if not integration_session:
                logger.warning("Webhook Uazapi ignorado; instancia desconhecida: %s", instance_identifier)
                continue
            debtor_phone: Optional[str] = None
            if is_group_message:
                phone_candidates = (
                    message_data.get("chatid"),
                    message_data.get("chatId"),
                    message_data.get("remoteJid"),
                    message_data.get("id"),
                )
            else:
                phone_candidates = (
                    message_data.get("sender_pn"),
                    message_data.get("chatid"),
                    message_data.get("chatId"),
                    message_data.get("remoteJid"),
                    message_data.get("sender"),
                )
            for candidate in phone_candidates:
                normalized_candidate = normalize_whatsapp_jid(candidate)
                if normalized_candidate:
                    debtor_phone = normalized_candidate
                    break
            if not debtor_phone:
                debtor_phone = normalize_phone(message_data.get("phone"))
            if not debtor_phone:
                continue
            message_type = detect_uazapi_message_type(message_data)
            text = extract_uazapi_text(message_data) or ""
            display_name = resolve_uazapi_debtor_name(message_data, chat_context)
            sender_display = (
                normalize_display_name(message_data.get("senderName"))
                or normalize_display_name(message_data.get("pushName"))
                or normalize_display_name(message_data.get("participant"))
                or normalize_display_name(message_data.get("sender"))
            )
            media_blob: Optional[bytes] = None
            media_content_type: Optional[str] = None
            media_original_name: Optional[str] = None
            media_duration = None
            duration_candidates = [
                message_data.get("seconds"),
                message_data.get("duration"),
                message_data.get("mediaDuration"),
            ]
            for candidate in duration_candidates:
                if isinstance(candidate, (int, float)):
                    media_duration = float(candidate)
                    break
            if message_type in {MessageType.audio, MessageType.image, MessageType.document}:
                downloaded = download_uazapi_media(message_data, integration_session)
                if downloaded:
                    media_blob, media_content_type, media_original_name = downloaded
                else:
                    if not text:
                        if message_type == MessageType.audio:
                            text = "[Audio recebido]"
                        else:
                            text = "[Arquivo recebido]"
                    message_type = MessageType.text
                    media_duration = None
            if is_group_message:
                label = sender_display or normalize_whatsapp_jid(
                    message_data.get("sender") or message_data.get("participant") or ""
                )
                if label:
                    if text:
                        text = f"{label}: {text}"
                    else:
                        text = f"{label}: [Mensagem recebida]"

            quoted_reference = extract_uazapi_reply_reference(message_data)
            incoming_payload = IncomingMessagePayload(
                debtor_phone=debtor_phone,
                debtor_name=display_name,
                content=text,
                session_code=integration_session.session_code,
                integration_message_id=message_data.get("messageid") or message_data.get("id"),
                message_type=message_type,
                media_blob=media_blob,
                media_original_name=media_original_name,
                media_content_type=media_content_type,
                media_duration_seconds=media_duration,
                quoted_integration_message_id=quoted_reference,
            )
            simulate_incoming_message(incoming_payload, session)
            processed += 1
    return {"processed": processed}


@app.get("/api/session", response_model=SessionRead)
def get_session_info(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> SessionRead:
    integration_session = ensure_session_exists(
        session, current_user.organization_id, current_user.id
    )
    return serialize_session(integration_session)


@app.post("/api/session/rotate", response_model=SessionRead)
def rotate_session(
    payload: SessionRotateRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> SessionRead:
    existing_session = ensure_session_exists(
        session, current_user.organization_id, current_user.id
    )
    organization = session.get(Organization, current_user.organization_id)
    org_server_url = (organization.uazapi_server_url or "").strip() if organization else ""
    base_url = (
        payload.integration_base_url
        or existing_session.integration_base_url
        or org_server_url
        or ""
    ).strip()
    instance_id = (payload.integration_instance_id or existing_session.integration_instance_id or "").strip()
    token = (payload.integration_token or existing_session.integration_token or "").strip()
    provider = "uazapi"
    auto_provision = bool(payload.auto_provision)
    provider_payload: Dict[str, Any] = {}
    instance_provisioned = False
    organization_credentials_available = bool(
        organization and organization.uazapi_server_url and organization.uazapi_admin_token
    )
    if provider == "uazapi" and not auto_provision and organization_credentials_available and not payload.integration_token:
        logger.info(
            "Auto-provisioning enabled using organization credentials org=%s", current_user.organization_id
        )
        auto_provision = True
    if provider == "uazapi" and not base_url and UAZAPI_DEFAULT_BASE_URL:
        base_url = UAZAPI_DEFAULT_BASE_URL

    print(
        "[rotate_session]",
        {"org": current_user.organization_id,
         "user": current_user.id,
         "base_provided": bool(payload.integration_base_url),
         "instance_provided": bool(payload.integration_instance_id),
         "token_provided": bool(payload.integration_token),
         "provider": provider,
         "auto_provision": auto_provision,
         }
    )
    webhook_id: Optional[str] = None
    admin_client: Optional[UazapiAdminClient] = None
    if provider == "uazapi" and auto_provision:
        admin_client = UazapiAdminClient(
            base_url=org_server_url or UAZAPI_ADMIN_BASE_URL,
            admin_token=(
                (organization.uazapi_admin_token or UAZAPI_ADMIN_TOKEN)
                if organization
                else UAZAPI_ADMIN_TOKEN
            ),
        )
        if not admin_client.is_configured:
            raise HTTPException(
                status_code=503,
                detail="Provisionamento automatico requer URL e admintoken da Uazapi definidos para a organização.",
            )
        org_seed = (
            organization.slug if isinstance(organization, Organization) else f"org{current_user.organization_id}"
        )
        desired_name = generate_uazapi_instance_name(payload.integration_instance_id, org_seed)
        try:
            creation_payload = admin_client.create_instance(
            name=desired_name,
            admin_field_01=f"org:{current_user.organization_id}",
            admin_field_02=f"user:{current_user.id}",
        )
            logger.info("Uazapi creation payload=%s", creation_payload)
        except Exception as exc:  # pragma: no cover - integracoes externas
            logger.error("Falha ao criar instancia na Uazapi: %s", exc)
            raise HTTPException(
                status_code=502,
                detail="Falha ao criar instancia automaticamente na Uazapi.",
            ) from exc
        instance_block = creation_payload.get("instance") or {}
        instance_id = instance_block.get("name") or creation_payload.get("name") or desired_name
        token = creation_payload.get("token") or instance_block.get("token") or token
        base_url = base_url or admin_client.base_url
        if not token or not instance_id or not base_url:
            raise HTTPException(
                status_code=502,
                detail="Uazapi nao retornou ID/token/base URL validos ao criar a instancia.",
            )
        instance_provisioned = True
        webhook_target_url = (
            (organization.webhook_url or "").strip() if organization else ""
        ) or WEBHOOK_URL
        logger.info("Webhook target url=%s", webhook_target_url)
        if webhook_target_url:
            logger.info("Skipping per-instance webhook configuration because webhook global is active.")

    if provider == "wppconnect":
        if not base_url or not instance_id:
            raise HTTPException(
                status_code=400,
                detail="Informe URL base e nome da sessao WPPConnect.",
            )
        qr_payload = (
            payload.qr_code_payload
            or existing_session.qr_code_payload
            or "Sessao WPPConnect configurada externamente."
        )
    elif provider == "uazapi":
        if not base_url:
            raise HTTPException(
                status_code=400,
                detail="Informe a URL da Uazapi ou configure UAZAPI_DEFAULT_BASE_URL.",
            )
        qr_payload = (
            None
            if auto_provision
            else payload.qr_code_payload or existing_session.qr_code_payload
        )
        if not instance_id or not token:
            raise HTTPException(
                status_code=400,
                detail="Informe ID da instancia e token da Uazapi.",
            )
        uazapi_client = UazapiClient(
            base_url=base_url,
            instance_id=instance_id,
            token=token,
        )

        def _extract_qr(data: Optional[Dict[str, Any]]) -> Optional[str]:
            if not isinstance(data, dict):
                return None
            instance_data = data.get("instance") or {}
            status_data = data.get("status") or {}
            for container in (instance_data, status_data, data):
                if isinstance(container, dict):
                    candidate = container.get("qrcode") or container.get("qrCode")
                    if isinstance(candidate, str) and candidate.strip():
                        return candidate.strip()
            return None

        def _extract_paircode(data: Optional[Dict[str, Any]]) -> Optional[str]:
            if not isinstance(data, dict):
                return None
            instance_data = data.get("instance") or {}
            candidate = instance_data.get("paircode") if isinstance(instance_data, dict) else None
            if isinstance(candidate, str) and candidate.strip():
                return candidate.strip()
            return None

        if not qr_payload:
            connect_payload: Optional[Dict[str, Any]] = None
            try:
                # Always regenerate QR by resetting any previous payload.
                existing_session.qr_code_payload = None
                session.add(existing_session)
                session.commit()
                connect_payload = uazapi_client.connect_instance()
                logger.info("Uazapi connect_instance payload=%s", connect_payload)
                provider_payload["connect"] = connect_payload
                qr_payload = _extract_qr(connect_payload)
            except Exception as exc:  # pragma: no cover - integrações externas
                logger.warning("Nao foi possivel iniciar conexao na Uazapi: %s", exc)
            if not qr_payload:
                status_payload: Optional[Dict[str, Any]] = None
                try:
                    status_payload = uazapi_client.fetch_instance_status()
                    logger.info("Uazapi fetch_instance_status payload=%s", status_payload)
                    provider_payload["status"] = status_payload
                    qr_payload = _extract_qr(status_payload)
                except Exception as exc:  # pragma: no cover - integrações externas
                    logger.warning("Nao foi possivel obter status da Uazapi: %s", exc)
                if not qr_payload:
                    paircode = _extract_paircode(connect_payload) or _extract_paircode(status_payload)
                    if paircode:
                        qr_payload = f"Codigo de pareamento: {paircode}"

        if not qr_payload:
            qr_payload = "Conecte o numero via painel Uazapi."
    else:
        candidate_client = EvolutionClient.from_values(
            base_url=base_url or None,
            instance_id=instance_id or None,
            token=token or None,
            admin_token=evolution_client.admin_token,
        )
        if not candidate_client.is_configured and evolution_client.is_configured:
            candidate_client = evolution_client
            base_url = evolution_client.base_url
            instance_id = evolution_client.instance_id
            token = evolution_client.token
            provider = "evolution"
        if not candidate_client.is_configured:
            raise HTTPException(
                status_code=400,
                detail="Informe URL base, ID da instancia e token da Evolution para gerar o QR.",
            )

        base_url = candidate_client.base_url
        instance_id = candidate_client.instance_id
        token = candidate_client.token

        qr_payload = payload.qr_code_payload or candidate_client.fetch_instance_qr()
        if not qr_payload:
            raise HTTPException(status_code=502, detail="Nao foi possivel obter o QR da Evolution")

    qr_payload_preview: Optional[str] = None
    qr_payload_length = 0
    if qr_payload:
        cleaned_payload = qr_payload.strip()
        qr_payload_length = len(cleaned_payload)
        qr_payload_preview = (
            cleaned_payload
            if qr_payload_length <= 60
            else f"{cleaned_payload[:60]}..."
        )
        qr_payload = cleaned_payload
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug(
                "rotate_session qr payload full len=%s data=%s",
                qr_payload_length,
                cleaned_payload,
            )
    logger.info(
        "rotate_session result org=%s user=%s provider=%s auto_provision=%s qr_len=%s qr_preview=%s payload=%s",
        current_user.organization_id,
        current_user.id,
        provider,
        auto_provision,
        qr_payload_length,
        qr_payload_preview,
        provider_payload,
    )
    existing_sessions = session.exec(
        select(DeviceSession)
        .where(DeviceSession.organization_id == current_user.organization_id)
        .where(DeviceSession.owner_user_id == current_user.id)
    ).all()
    for previous in existing_sessions:
        previous.status = DeviceSessionStatus.disconnected
        previous.qr_code_payload = None
        session.add(previous)

    integration_session = DeviceSession(
        organization_id=current_user.organization_id,
        owner_user_id=current_user.id,
        session_label=payload.session_label or "primary",
        qr_code_payload=qr_payload,
        status=DeviceSessionStatus.pending,
        provider=provider or "uazapi",
        integration_base_url=base_url,
        integration_instance_id=instance_id,
        integration_token=token,
        webhook_id=webhook_id,
    )
    session.add(integration_session)
    session.commit()
    session.refresh(integration_session)
    if instance_provisioned:
        logger.info(
            "Uazapi instancia provisionada para org=%s user=%s instance=%s provider=%s",
            current_user.organization_id,
            current_user.id,
            instance_id,
            provider,
        )
    return serialize_session(integration_session)


@app.post("/api/auth/login", response_model=AuthResponse)
def login(payload: AuthLogin, session: Session = Depends(get_session)) -> AuthResponse:
    logger.debug("login payload: %s", payload.model_dump())
    username = normalize_username(payload.username)
    organization = None
    if payload.collaborator:
        user = session.exec(
            select(User)
            .where(User.username == username)
            .where(User.organization_id == None)  # noqa: E711
        ).first()
        if not user or not verify_password(payload.password, user.password_hash):
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Credenciais invalidas")
    else:
        if not payload.organization_slug or not payload.organization_slug.strip():
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Slug obrigatorio")
        organization_slug = slugify(payload.organization_slug)
        organization = session.exec(
            select(Organization).where(Organization.slug == organization_slug)
        ).first()
        if not organization:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Organizacao desconhecida")
        user = session.exec(
            select(User)
            .where(User.username == username)
            .where(User.organization_id == organization.id)
        ).first()
        if not user or not verify_password(payload.password, user.password_hash):
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Credenciais invalidas")
    if not user.is_active:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Usuario inativo")
    token = create_session_token(user, session)
    return AuthResponse(
        token=token.token,
        user=UserRead.model_validate(user),
        organization=build_organization_read(organization),
    )


@app.post("/api/auth/register", response_model=AuthResponse)
def register(payload: AuthRegister, session: Session = Depends(get_session)) -> AuthResponse:
    slug = slugify(payload.organization_slug)
    username = normalize_username(payload.username)
    existing = session.exec(select(Organization).where(Organization.slug == slug)).first()
    if existing:
        raise HTTPException(status_code=409, detail="Organizacao ja existe")
    organization = Organization(name=payload.organization_name, slug=slug)
    session.add(organization)
    session.commit()
    session.refresh(organization)

    try:
        user = User(
            username=username,
            full_name=payload.full_name,
            password_hash=get_password_hash(payload.password),
            organization_id=organization.id,
            is_admin=True,
        )
        session.add(user)
        session.commit()
        session.refresh(user)
    except IntegrityError:
        session.rollback()
        raise HTTPException(status_code=409, detail="Usuario ja existe")

    token = create_session_token(user, session)
    return AuthResponse(
        token=token.token,
        user=UserRead.model_validate(user),
        organization=build_organization_read(organization),
    )


@app.post("/api/auth/logout")
def logout(
    session: Session = Depends(get_session),
    authorization: Optional[str] = Header(default=None),
) -> dict:
    token_value = extract_bearer_token(authorization)
    token = session.exec(select(SessionToken).where(SessionToken.token == token_value)).first()
    if token:
        session.delete(token)
        session.commit()
    return {"status": "ok"}


@app.get("/api/auth/me")
def read_current_user(current_user: User = Depends(get_current_user)) -> dict:
    organization = current_user.organization if current_user.organization else None
    return {
        "user": UserRead.model_validate(current_user),
        "organization": OrganizationRead.model_validate(organization)
        if organization
        else None,
    }


@app.get("/api/organization", response_model=OrganizationRead)
def read_current_organization(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> OrganizationRead:
    organization = session.get(Organization, current_user.organization_id)
    if not organization:
        raise HTTPException(status_code=404, detail="Organizacao nao encontrada.")
    result = build_organization_read(organization)
    if not result:
        raise HTTPException(status_code=404, detail="Organizacao nao encontrada.")
    return result


@app.get("/api/users", response_model=List[UserRead])
def list_users(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> List[UserRead]:
    require_admin(current_user)
    users = session.exec(
        select(User)
        .where(User.organization_id == current_user.organization_id)
        .order_by(User.created_at.asc())
    ).all()
    user_ids = [user.id for user in users if user.id]
    dept_map: Dict[int, List[str]] = {}
    if user_ids:
        dept_rows = session.exec(
            select(DepartmentUserLink.user_id, Department.name)
            .where(DepartmentUserLink.user_id.in_(user_ids))
            .where(DepartmentUserLink.department_id == Department.id)
            .where(Department.organization_id == current_user.organization_id)
        ).all()
        for user_id, dept_name in dept_rows:
            if not user_id or not dept_name:
                continue
            dept_map.setdefault(int(user_id), []).append(str(dept_name))
        for user_id, names in dept_map.items():
            dept_map[user_id] = sorted({n.strip() for n in names if n.strip()})

    result: List[UserRead] = []
    for user in users:
        data = UserRead.model_validate(user)
        data.departments = dept_map.get(user.id or 0, [])
        result.append(data)
    return result


@app.get("/api/departments", response_model=List[DepartmentRead])
def list_departments(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> List[DepartmentRead]:
    require_admin(current_user)
    departments = session.exec(
        select(Department)
        .where(Department.organization_id == current_user.organization_id)
        .order_by(Department.created_at.asc())
    ).all()
    dept_ids = [dept.id for dept in departments if dept.id]
    counts: Dict[int, int] = {}
    if dept_ids:
        rows = session.exec(
            select(DepartmentUserLink.department_id, func.count(DepartmentUserLink.user_id))
            .where(DepartmentUserLink.department_id.in_(dept_ids))
            .group_by(DepartmentUserLink.department_id)
        ).all()
        counts = {dept_id: count for dept_id, count in rows if dept_id}
    return [
        DepartmentRead(
            id=dept.id,
            organization_id=dept.organization_id,
            name=dept.name,
            created_at=dept.created_at,
            member_count=counts.get(dept.id or 0, 0),
        )
        for dept in departments
        if dept.id is not None
    ]


@app.post("/api/departments", response_model=DepartmentRead, status_code=201)
def create_department(
    payload: DepartmentCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> DepartmentRead:
    require_admin(current_user)
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=422, detail="Informe o nome do departamento.")
    existing = session.exec(
        select(Department).where(
            Department.organization_id == current_user.organization_id,
            func.lower(Department.name) == name.lower(),
        )
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail="Departamento já existe.")

    dept = Department(organization_id=current_user.organization_id, name=name)
    session.add(dept)
    session.commit()
    session.refresh(dept)
    return DepartmentRead(
        id=dept.id,
        organization_id=dept.organization_id,
        name=dept.name,
        created_at=dept.created_at,
        member_count=0,
    )


@app.patch("/api/departments/{department_id}", response_model=DepartmentRead)
def update_department(
    department_id: int,
    payload: DepartmentUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> DepartmentRead:
    require_admin(current_user)
    dept = session.get(Department, department_id)
    if not dept or dept.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Departamento não encontrado.")
    update_data = payload.model_dump(exclude_unset=True)
    if "name" in update_data:
        name = (update_data["name"] or "").strip()
        if not name:
            raise HTTPException(status_code=422, detail="Informe o nome do departamento.")
        existing = session.exec(
            select(Department).where(
                Department.organization_id == current_user.organization_id,
                func.lower(Department.name) == name.lower(),
                Department.id != department_id,
            )
        ).first()
        if existing:
            raise HTTPException(status_code=409, detail="Já existe um departamento com esse nome.")
        dept.name = name
    session.add(dept)
    session.commit()
    session.refresh(dept)
    member_count = session.exec(
        select(func.count(DepartmentUserLink.user_id)).where(
            DepartmentUserLink.department_id == department_id
        )
    ).one()
    return DepartmentRead(
        id=dept.id,
        organization_id=dept.organization_id,
        name=dept.name,
        created_at=dept.created_at,
        member_count=int(member_count or 0),
    )


@app.delete("/api/departments/{department_id}", status_code=204)
def delete_department(
    department_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    require_admin(current_user)
    dept = session.get(Department, department_id)
    if not dept or dept.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Departamento não encontrado.")
    session.exec(delete(DepartmentUserLink).where(DepartmentUserLink.department_id == department_id))
    session.delete(dept)
    session.commit()


@app.get("/api/departments/{department_id}/users", response_model=List[int])
def list_department_users(
    department_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> List[int]:
    require_admin(current_user)
    dept = session.get(Department, department_id)
    if not dept or dept.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Departamento não encontrado.")
    rows = session.exec(
        select(DepartmentUserLink.user_id).where(DepartmentUserLink.department_id == department_id)
    ).all()
    return [int(user_id) for user_id in rows if user_id is not None]


@app.put("/api/departments/{department_id}/users", status_code=204)
def set_department_users(
    department_id: int,
    payload: DepartmentUsersUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    require_admin(current_user)
    dept = session.get(Department, department_id)
    if not dept or dept.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Departamento não encontrado.")

    user_ids = list({int(uid) for uid in (payload.user_ids or []) if int(uid) > 0})
    if user_ids:
        allowed = session.exec(
            select(User.id).where(
                User.organization_id == current_user.organization_id,
                User.id.in_(user_ids),
            )
        ).all()
        allowed_set = {int(uid) for uid in allowed if uid is not None}
        if allowed_set != set(user_ids):
            raise HTTPException(status_code=400, detail="Um ou mais usuários são inválidos.")

    session.exec(delete(DepartmentUserLink).where(DepartmentUserLink.department_id == department_id))
    for user_id in user_ids:
        session.add(DepartmentUserLink(department_id=department_id, user_id=user_id))
    session.commit()


@app.post("/api/users", response_model=UserRead, status_code=201)
def create_user(
    payload: UserCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> UserRead:
    require_admin(current_user)
    username = normalize_username(payload.username)
    existing = session.exec(select(User).where(User.username == username)).first()
    if existing:
        raise HTTPException(status_code=409, detail="Usuario ja existe")
    user = User(
        username=username,
        full_name=payload.full_name,
        password_hash=get_password_hash(payload.password),
        organization_id=current_user.organization_id,
        is_admin=payload.is_admin,
    )
    session.add(user)
    session.commit()
    session.refresh(user)
    return UserRead.model_validate(user)


@app.get("/api/admin/uazapi-credentials", response_model=OrganizationCredentialsRead)
def read_organization_uazapi_credentials(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> OrganizationCredentialsRead:
    ensure_admin_access(current_user)
    organization = session.get(Organization, current_user.organization_id)
    if not organization:
        raise HTTPException(status_code=404, detail="Organizacao nao encontrada.")
    return OrganizationCredentialsRead(
        uazapi_server_url=organization.uazapi_server_url,
        has_admin_token=bool(organization.uazapi_admin_token),
    )


@app.put("/api/admin/uazapi-credentials", response_model=OrganizationCredentialsRead)
def update_organization_uazapi_credentials(
    payload: OrganizationCredentialsUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> OrganizationCredentialsRead:
    ensure_admin_access(current_user)
    organization = session.get(Organization, current_user.organization_id)
    if not organization:
        raise HTTPException(status_code=404, detail="Organizacao nao encontrada.")
    if payload.uazapi_server_url is not None:
        organization.uazapi_server_url = payload.uazapi_server_url.strip() or None
    if payload.uazapi_admin_token is not None:
        organization.uazapi_admin_token = payload.uazapi_admin_token.strip() or None
    session.add(organization)
    session.commit()
    session.refresh(organization)
    return OrganizationCredentialsRead(
        uazapi_server_url=organization.uazapi_server_url,
        has_admin_token=bool(organization.uazapi_admin_token),
    )


@app.get("/api/admin/organizations", response_model=List[CollaboratorOrganizationRead])
def list_admin_organizations(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> List[CollaboratorOrganizationRead]:
    ensure_admin_access(current_user)
    organizations = session.exec(select(Organization).order_by(Organization.created_at.desc())).all()
    result: List[CollaboratorOrganizationRead] = []
    for organization in organizations:
        user_count = get_organization_user_count(session, organization.id)
        admin_username = get_primary_admin_username(session, organization.id)
        result.append(
            serialize_collaborator_organization(organization, user_count, admin_username)
        )
    return result


@app.post(
    "/api/admin/organizations",
    response_model=CollaboratorOrganizationRead,
    status_code=201,
)
def create_admin_organization(
    payload: CollaboratorOrganizationCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> CollaboratorOrganizationRead:
    ensure_admin_access(current_user)
    if not payload.slug.strip():
        raise HTTPException(status_code=422, detail="Slug obrigatorio.")
    slug = slugify(payload.slug)
    existing = session.exec(select(Organization).where(Organization.slug == slug)).first()
    if existing:
        raise HTTPException(status_code=409, detail="Organizacao ja existe.")
    organization = Organization(
        name=payload.name.strip(),
        slug=slug,
        uazapi_server_url=payload.uazapi_server_url.strip()
        if payload.uazapi_server_url
        else None,
        uazapi_admin_token=payload.uazapi_admin_token.strip()
        if payload.uazapi_admin_token
        else None,
        webhook_url=payload.webhook_url.strip() if payload.webhook_url else None,
        user_limit=max(0, payload.user_limit),
    )
    session.add(organization)
    session.commit()
    session.refresh(organization)
    try:
        admin_username = normalize_username(payload.admin_username)
        user = User(
            username=admin_username,
            full_name=payload.admin_full_name,
            password_hash=get_password_hash(payload.admin_password),
            organization_id=organization.id,
            is_admin=True,
        )
        session.add(user)
        session.commit()
    except IntegrityError:
        session.rollback()
        session.delete(organization)
        session.commit()
        raise HTTPException(status_code=409, detail="Usuario administrador ja existe.")
    finally:
        session.refresh(organization)
    user_count = get_organization_user_count(session, organization.id)
    admin_username = get_primary_admin_username(session, organization.id)
    return serialize_collaborator_organization(organization, user_count, admin_username)


@app.get(
    "/api/admin/organizations/{organization_id}/credentials",
    response_model=OrganizationCredentialsRead,
)
def read_collaborator_organization_credentials(
    organization_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> OrganizationCredentialsRead:
    ensure_admin_access(current_user)
    organization = session.get(Organization, organization_id)
    if not organization:
        raise HTTPException(status_code=404, detail="Organizacao nao encontrada.")
    return OrganizationCredentialsRead(
        uazapi_server_url=organization.uazapi_server_url,
        has_admin_token=bool(organization.uazapi_admin_token),
    )


@app.put(
    "/api/admin/organizations/{organization_id}/credentials",
    response_model=OrganizationCredentialsRead,
)
def update_collaborator_organization_credentials(
    organization_id: int,
    payload: OrganizationCredentialsUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> OrganizationCredentialsRead:
    ensure_admin_access(current_user)
    organization = session.get(Organization, organization_id)
    if not organization:
        raise HTTPException(status_code=404, detail="Organizacao nao encontrada.")
    if payload.uazapi_server_url is not None:
        organization.uazapi_server_url = payload.uazapi_server_url.strip() or None
    if payload.uazapi_admin_token is not None:
        organization.uazapi_admin_token = payload.uazapi_admin_token.strip() or None
    if payload.webhook_url is not None:
        organization.webhook_url = payload.webhook_url.strip() or None
    session.add(organization)
    session.commit()
    session.refresh(organization)
    return OrganizationCredentialsRead(
        uazapi_server_url=organization.uazapi_server_url,
        has_admin_token=bool(organization.uazapi_admin_token),
    )


@app.patch(
    "/api/admin/organizations/{organization_id}",
    response_model=CollaboratorOrganizationRead,
)
def update_admin_organization(
    organization_id: int,
    payload: OrganizationUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> CollaboratorOrganizationRead:
    ensure_admin_access(current_user)
    organization = session.get(Organization, organization_id)
    if not organization:
        raise HTTPException(status_code=404, detail="Organizacao nao encontrada.")
    if payload.is_active is not None:
        organization.is_active = payload.is_active
    if payload.user_limit is not None:
        if payload.user_limit < 0:
            raise HTTPException(status_code=422, detail="Limite de usuarios invalido.")
        organization.user_limit = payload.user_limit
    if payload.webhook_url is not None:
        organization.webhook_url = payload.webhook_url.strip() or None
    session.add(organization)
    session.commit()
    session.refresh(organization)
    user_count = get_organization_user_count(session, organization.id)
    admin_username = get_primary_admin_username(session, organization.id)
    return serialize_collaborator_organization(organization, user_count, admin_username)


@app.get("/api/tags", response_model=List[TagRead])
def list_tags(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> List[TagRead]:
    tags = session.exec(
        select(Tag)
        .where(Tag.organization_id == current_user.organization_id)
        .where(
            or_(
                Tag.owner_user_id == None,  # noqa: E711
                Tag.owner_user_id == current_user.id,
            )
        )
        .order_by(Tag.name.asc())
    ).all()
    return [TagRead.model_validate(tag) for tag in tags]


@app.post("/api/tags", response_model=TagRead, status_code=201)
def create_tag(
    payload: TagCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> TagRead:
    tag = Tag(
        name=payload.name,
        color=payload.color,
        organization_id=current_user.organization_id,
        owner_user_id=current_user.id,
    )
    session.add(tag)
    session.commit()
    session.refresh(tag)
    return TagRead.model_validate(tag)


@app.delete("/api/tags/{tag_id}", status_code=204)
def delete_tag(
    tag_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    tag = session.get(Tag, tag_id)
    if (
        not tag
        or tag.organization_id != current_user.organization_id
        or (tag.owner_user_id not in (None, current_user.id))
    ):
        raise HTTPException(status_code=404, detail="Etiqueta não encontrada")
    session.exec(
        delete(ConversationTagLink).where(ConversationTagLink.tag_id == tag_id)
    )
    session.delete(tag)
    session.commit()


@app.post("/api/conversations/{conversation_id}/tags/{tag_id}", status_code=204)
def attach_tag(
    conversation_id: int,
    tag_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    conversation = session.get(Conversation, conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa não encontrada")
    ensure_conversation_access(conversation, current_user)
    tag = session.get(Tag, tag_id)
    if (
        not tag
        or tag.organization_id != current_user.organization_id
        or (tag.owner_user_id not in (None, current_user.id))
    ):
        raise HTTPException(status_code=404, detail="Etiqueta não encontrada")
    existing = session.exec(
        select(ConversationTagLink)
        .where(ConversationTagLink.conversation_id == conversation_id)
        .where(ConversationTagLink.tag_id == tag_id)
    ).first()
    if not existing:
        link = ConversationTagLink(
            conversation_id=conversation_id, tag_id=tag_id
        )
        session.add(link)
        session.commit()


@app.delete("/api/conversations/{conversation_id}/tags/{tag_id}", status_code=204)
def detach_tag(
    conversation_id: int,
    tag_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    conversation = session.get(Conversation, conversation_id)
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversa não encontrada")
    ensure_conversation_access(conversation, current_user)
    tag = session.get(Tag, tag_id)
    if (
        not tag
        or tag.organization_id != current_user.organization_id
        or (tag.owner_user_id not in (None, current_user.id))
    ):
        raise HTTPException(status_code=404, detail="Etiqueta não encontrada")
    session.exec(
        delete(ConversationTagLink)
        .where(ConversationTagLink.conversation_id == conversation_id)
        .where(ConversationTagLink.tag_id == tag_id)
    )
    session.commit()


@app.get("/api/bulk/campaigns", response_model=List[BulkCampaignRead])
def list_bulk_campaigns(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> List[BulkCampaignRead]:
    campaigns = session.exec(
        select(BulkCampaign)
        .where(BulkCampaign.organization_id == current_user.organization_id)
        .order_by(BulkCampaign.created_at.desc())
    ).all()
    maybe_sync_bulk_campaigns(session, current_user, campaigns)
    return [BulkCampaignRead.model_validate(campaign) for campaign in campaigns]


@app.post("/api/bulk/campaigns", response_model=BulkCampaignRead)
async def create_bulk_campaign(
    campaign: str = Form(..., description="JSON contendo os dados da campanha"),
    contacts_file: UploadFile = File(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> BulkCampaignRead:
    try:
        payload = json.loads(campaign)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="JSON invalido para a campanha.") from exc
    try:
        data = BulkCampaignCreate.model_validate(payload)
    except ValidationError as exc:
        raise HTTPException(status_code=422, detail=exc.errors())
    file_blob = await contacts_file.read()
    if not file_blob:
        raise HTTPException(status_code=400, detail="Envie a planilha de contatos.")
    if len(file_blob) > MAX_UPLOAD_SIZE_BYTES:
        raise HTTPException(status_code=400, detail="Arquivo deve ter até 10 MB.")
    _, rows = extract_rows_from_file(file_blob, contacts_file.filename or "")
    contacts = parse_contacts_from_rows(rows, data.mapping)
    template = (data.message_template or "").strip()
    if not template:
        raise HTTPException(status_code=400, detail="Informe o texto da mensagem.")
    if len(contacts) > MAX_CAMPAIGN_CONTACTS:
        raise HTTPException(
            status_code=400,
            detail=f"Limite de {MAX_CAMPAIGN_CONTACTS} contatos por campanha.",
        )
    messages: List[Dict[str, Any]] = []
    local_message_records: List[Tuple[Conversation, str]] = []
    conversation_cache: Dict[str, Conversation] = {}
    for contact in contacts:
        context = build_template_context(contact)
        text = render_message_template(template, context).strip()
        if not text:
            continue
        messages.append({"number": contact["phone"], "type": "text", "text": text})
        cached_conversation = conversation_cache.get(contact["phone"])
        if not cached_conversation:
            cached_conversation = ensure_conversation_for_phone(
                session,
                current_user.organization_id,
                contact["phone"],
                display_name=contact.get("name") or contact.get("raw_phone"),
                owner_user_id=current_user.id,
            )
            conversation_cache[contact["phone"]] = cached_conversation
        local_message_records.append((cached_conversation, text))
    if not messages:
        raise HTTPException(
            status_code=400,
            detail="Nenhuma mensagem valida foi gerada a partir da planilha.",
        )
    integration_session = ensure_session_exists(
        session, current_user.organization_id, current_user.id
    )
    client = get_gateway_client(integration_session)
    if not isinstance(client, UazapiClient) or not client.is_configured:
        raise HTTPException(
            status_code=400,
            detail="Configure o canal WhatsApp Nexen na aba Integracao antes de criar campanhas.",
        )
    scheduled_dt = data.scheduled_for
    scheduled_timestamp: Optional[int] = None
    if scheduled_dt and scheduled_dt > datetime.utcnow():
        scheduled_timestamp = int(scheduled_dt.timestamp() * 1000)
    else:
        scheduled_dt = None
    status_value = "scheduled" if scheduled_timestamp else "running"
    folder_label = slugify(data.name)
    if not folder_label:
        folder_label = f"campanha-{uuid4().hex[:6]}"
    delay_value = max(1, data.send_interval_seconds or 3)
    request_payload: Dict[str, Any] = {
        "folder": folder_label,
        "info": data.preview_text or data.name,
        "messages": messages,
        "delayMin": delay_value,
        "delayMax": delay_value,
    }
    if scheduled_timestamp:
        request_payload["scheduled_for"] = scheduled_timestamp
    try:
        provider_response = client.send_mass_campaign(request_payload)
    except requests.HTTPError as exc:  # pragma: no cover - integracao externa
        body = ""
        if getattr(exc, "response", None) is not None:
            try:
                body = exc.response.text[:500]  # type: ignore[attr-defined]
            except Exception:
                body = ""
        logger.error("Falha ao criar campanha na Uazapi: %s - body=%s", exc, body)
        detail = body or "Falha ao enviar campanha para a Uazapi."
        raise HTTPException(status_code=502, detail=detail) from exc
    except Exception as exc:  # pragma: no cover - integracao externa
        logger.error("Falha ao criar campanha na Uazapi: %s", exc)
        raise HTTPException(status_code=502, detail="Falha ao enviar campanha para a Uazapi.") from exc

    for conversation, text in local_message_records:
        message = Message(
            conversation_id=conversation.id,
            direction=MessageDirection.agent,
            content=text,
            is_read=True,
            author_user_id=current_user.id,
            message_type=MessageType.text,
            via_session_id=integration_session.id,
        )
        session.add(message)
        conversation.updated_at = datetime.utcnow()
        session.add(conversation)
    session.commit()

    folder_id = str(
        provider_response.get("folder_id")
        or provider_response.get("folderId")
        or provider_response.get("id")
        or folder_label
    )
    queued_count = extract_int(provider_response, "count", "queued", "total") or len(messages)
    campaign_record = BulkCampaign(
        organization_id=current_user.organization_id,
        created_by_user_id=current_user.id,
        name=data.name.strip(),
        channel=(data.channel or "WhatsApp Nexen").strip(),
        status=status_value,
        folder_id=folder_id,
        contacts_total=len(messages),
        sent_count=queued_count if status_value == "running" else 0,
        message_template=template,
        preview_text=data.preview_text,
        cost_estimate=data.cost_estimate,
        send_interval_seconds=data.send_interval_seconds,
        scheduled_for=scheduled_dt,
        started_at=datetime.utcnow() if status_value == "running" else None,
        mapping_schema=json.dumps(data.mapping.model_dump()),
    )
    session.add(campaign_record)
    session.commit()
    session.refresh(campaign_record)
    return BulkCampaignRead.model_validate(campaign_record)


@app.get("/api/reminders", response_model=List[ReminderRead])
def list_reminders(
    conversation_id: Optional[int] = Query(None),
    status: str = Query("pending", regex="^(pending|done|all)$"),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> List[ReminderRead]:
    query = select(Reminder).where(
        Reminder.organization_id == current_user.organization_id
    )
    if conversation_id:
        conversation = session.get(Conversation, conversation_id)
        if not conversation:
            raise HTTPException(status_code=404, detail="Conversa nao encontrada")
        ensure_conversation_access(conversation, current_user)
        query = query.where(Reminder.conversation_id == conversation_id)
    query = query.where(Reminder.owner_user_id == current_user.id)
    if status == "pending":
        query = query.where(Reminder.is_done == False)  # noqa: E712
    elif status == "done":
        query = query.where(Reminder.is_done == True)  # noqa: E712
    reminders = session.exec(
        query.order_by(Reminder.due_at.asc(), Reminder.id.asc())
    ).all()
    return [ReminderRead.model_validate(reminder) for reminder in reminders]


@app.post("/api/reminders", response_model=ReminderRead, status_code=201)
def create_reminder(
    payload: ReminderCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ReminderRead:
    conversation_id = payload.conversation_id
    if conversation_id:
        conversation = session.get(Conversation, conversation_id)
        if not conversation:
            raise HTTPException(status_code=404, detail="Conversa nao encontrada")
        ensure_conversation_access(conversation, current_user)
    owner_user_id = payload.owner_user_id or current_user.id
    if owner_user_id != current_user.id and not current_user.is_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Apenas administradores podem atribuir lembretes para outros usuarios",
        )
    owner_user = get_user_in_organization(
        session, owner_user_id, current_user.organization_id
    )
    if not owner_user:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    title = payload.title.strip()
    if not title:
        raise HTTPException(status_code=422, detail="Titulo obrigatorio")
    reminder = Reminder(
        title=title,
        due_at=payload.due_at,
        conversation_id=conversation_id,
        organization_id=current_user.organization_id,
        owner_user_id=owner_user.id,
        created_by_user_id=current_user.id,
    )
    session.add(reminder)
    session.commit()
    session.refresh(reminder)
    return ReminderRead.model_validate(reminder)


@app.patch("/api/reminders/{reminder_id}", response_model=ReminderRead)
def update_reminder(
    reminder_id: int,
    payload: ReminderUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ReminderRead:
    reminder = session.get(Reminder, reminder_id)
    if not reminder:
        raise HTTPException(status_code=404, detail="Lembrete nao encontrado")
    ensure_reminder_access(reminder, current_user)
    update_data = payload.model_dump(exclude_unset=True)
    if "title" in update_data:
        title = (update_data["title"] or "").strip()
        if not title:
            raise HTTPException(status_code=422, detail="Titulo obrigatorio")
        reminder.title = title
    if "due_at" in update_data and update_data["due_at"]:
        reminder.due_at = update_data["due_at"]
    if "is_done" in update_data and update_data["is_done"] is not None:
        reminder.is_done = update_data["is_done"]
    session.add(reminder)
    session.commit()
    session.refresh(reminder)
    return ReminderRead.model_validate(reminder)


@app.delete("/api/reminders/{reminder_id}", status_code=204)
def delete_reminder(
    reminder_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    reminder = session.get(Reminder, reminder_id)
    if not reminder:
        raise HTTPException(status_code=404, detail="Lembrete nao encontrado")
    ensure_reminder_access(reminder, current_user)
    session.delete(reminder)
    session.commit()


@app.patch("/api/users/{user_id}", response_model=UserRead)
def update_user(
    user_id: int,
    payload: UserUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> UserRead:
    require_admin(current_user)
    user = session.get(User, user_id)
    if not user or user.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    update_data = payload.model_dump(exclude_unset=True)
    if "full_name" in update_data:
        user.full_name = update_data["full_name"]
    if "is_active" in update_data and user.id != current_user.id:
        user.is_active = update_data["is_active"]
    if "password" in update_data and update_data["password"]:
        user.password_hash = get_password_hash(update_data["password"])
    if "is_admin" in update_data:
        new_is_admin = update_data["is_admin"]
        if user.id == current_user.id and not new_is_admin:
            raise HTTPException(status_code=400, detail="Nao eh permitido remover seu proprio perfil administrador")
        if not new_is_admin and not has_other_admin(session, user.organization_id, exclude_user_id=user.id):
            raise HTTPException(status_code=400, detail="Organizacao precisa de pelo menos um administrador")
        user.is_admin = new_is_admin
    session.add(user)
    session.commit()
    session.refresh(user)
    return UserRead.model_validate(user)
